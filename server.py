# server.py
from __future__ import annotations

import os
import io
import re
import json
import secrets
import unicodedata
from openpyxl import load_workbook
from flask import send_file
from sqlalchemy import delete, extract, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func
from datetime import datetime, timezone, timedelta
from datetime import date as _date
from datetime import date
from sqlalchemy.exc import IntegrityError
from calendar import monthrange
import pdfplumber
from dotenv import load_dotenv
from collections import Counter
from flask import Flask, request, jsonify, render_template, send_from_directory, redirect
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt, get_jwt_identity
from sqlalchemy import and_, text as sqltext
from sqlalchemy import text
from sqlalchemy.sql import func
from sqlalchemy.pool import NullPool
from sqlalchemy.exc import OperationalError, DisconnectionError
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.orm import sessionmaker
from flask import current_app


# ---------------------------------
# Config
# ---------------------------------
load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')

db_url = os.getenv('DATABASE_URL')
if not db_url:
    db_url = 'sqlite:///schedule.db'
elif db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql+psycopg2://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'dev-secret-change-me')

# Пул/антипул
engine_opts = {'pool_pre_ping': True}
if ('pooler.supabase.com' in db_url) or os.getenv('DB_FORCE_NULLPOOL') == '1':
    engine_opts['poolclass'] = NullPool
else:
    engine_opts.update({
        'pool_recycle': int(os.getenv('DB_POOL_RECYCLE', '300')),
        'pool_size': int(os.getenv('DB_POOL_SIZE', '3')),
        'max_overflow': int(os.getenv('DB_MAX_OVERFLOW', '0')),
    })
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = engine_opts

jwt = JWTManager(app)
db = SQLAlchemy(app)

ADMIN_EMAILS = {e.strip().lower() for e in os.getenv('ADMIN_EMAILS', '').split(',') if e.strip()}

# Кто может финально утверждать замены (нижний регистр!)
MANAGER_EMAILS = {"r.czajka@lot.pl", "m.kaczmarski@lot.pl"}

# ---------------------------------
# Models
# ---------------------------------
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=True)
    password_hash = db.Column(db.String(255), nullable=True)
    full_name = db.Column(db.String(255), unique=True, nullable=False)
    role = db.Column(db.String(50), nullable=False, default='user')
    order_index = db.Column(db.Integer, nullable=True)
    is_coordinator = db.Column(db.Boolean, nullable=False, default=False)
    is_zmiwaka = db.Column(db.Boolean, nullable=False, default=False)
    hourly_rate_pln = db.Column(db.Numeric(10, 2), nullable=True)
    tax_percent = db.Column(db.Numeric(5, 2), nullable=True, default=0)
    reset_token   = db.Column(db.String(128), nullable=True)
    reset_expires = db.Column(db.DateTime, nullable=True)


    shifts = db.relationship('Shift', backref='user', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'email': self.email,
            'full_name': self.full_name,
            'role': self.role,
            'order_index': self.order_index,
            'is_coordinator': bool(self.is_coordinator),
            'is_zmiwaka': bool(self.is_zmiwaka),
            'hourly_rate_pln': float(self.hourly_rate_pln) if self.hourly_rate_pln is not None else None,
            'tax_percent': float(self.tax_percent or 0),
        }


class Shift(db.Model):
    __tablename__ = 'shifts'
    __table_args__ = (
        db.UniqueConstraint('user_id', 'shift_date', name='uq_shifts_user_date'),
    )
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    shift_date = db.Column(db.Date, nullable=False)
    shift_code = db.Column(db.String(50), nullable=False)
    hours = db.Column(db.Integer, nullable=True)
    worked_hours = db.Column(db.Numeric(5, 2), nullable=True)
    work_note = db.Column(db.Text, nullable=True)

    # ↓↓↓ ДОБАВИТЬ
    lounge = db.Column(db.String(16))         # 'mazurek' | 'polonez' | None
    coord_lounge = db.Column(db.String(16))   # 'mazurek' | 'polonez' | None

    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'full_name': self.user.full_name if self.user else None,
            'shift_date': self.shift_date.isoformat(),
            'shift_code': self.shift_code,
            'hours': self.hours,
            'worked_hours': float(self.worked_hours) if self.worked_hours is not None else None,
            'lounge': self.lounge,
            'lounge': self.lounge,
            'coord_lounge': self.coord_lounge,
        }

# --- модель контроля (рядом с другими моделями) ---
class ControlEvent(db.Model):
    __tablename__ = 'control_events'
    id            = db.Column(db.Integer, primary_key=True)
    kind          = db.Column(db.String(20), nullable=False)  # 'late' | 'extra' | 'absence' | 'manual_shift'
    user_id       = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    event_date    = db.Column(db.Date, nullable=False)
    reason        = db.Column(db.Text, nullable=True)
    hours         = db.Column(db.Numeric(5,2), nullable=True)  # для extra
    time_from     = db.Column(db.String(5), nullable=True)     # HH:MM (manual_shift)
    time_to       = db.Column(db.String(5), nullable=True)     # HH:MM (manual_shift)
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at    = db.Column(db.DateTime(timezone=True), server_default=func.now())

    user         = db.relationship('User', foreign_keys=[user_id])
    created_by   = db.relationship('User', foreign_keys=[created_by_id])

    def to_dict(self):
        return {
            'id': self.id, 'kind': self.kind, 'user_id': self.user_id,
            'user': self.user.full_name if self.user else None,
            'date': self.event_date.isoformat(),
            'reason': self.reason,
            'hours': float(self.hours) if self.hours is not None else None,
            'time_from': self.time_from, 'time_to': self.time_to,
            'created_by': self.created_by.full_name if self.created_by else None,
            'created_at': self.created_at.isoformat() if self.created_at else None,
        }


class SwapProposal(db.Model):
    __tablename__ = 'swap_proposals'
    id = db.Column(db.Integer, primary_key=True)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    target_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    my_date = db.Column(db.Date, nullable=False)      # моё (что отдаю)
    their_date = db.Column(db.Date, nullable=False)   # их (что хочу)
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending/accepted/declined/canceled/approved/rejected
    created_at = db.Column(db.DateTime(timezone=True), server_default=func.now())

    requester = db.relationship('User', foreign_keys=[requester_id])
    target_user = db.relationship('User', foreign_keys=[target_user_id])

    def to_dict(self):
        return {
            'id': self.id,
            'requester': self.requester.to_dict() if self.requester else None,
            'target_user': self.target_user.to_dict() if self.target_user else None,
            'my_date': self.my_date.isoformat(),
            'their_date': self.their_date.isoformat(),
            'status': self.status,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class MarketOffer(db.Model):
    __tablename__ = 'market_offers'
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shifts.id'), nullable=False, unique=True)
    owner_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    status = db.Column(db.String(20), nullable=False, default='open')  # open/requested/approved/rejected/cancelled
    created_at = db.Column(db.DateTime(timezone=True), server_default=func.now())

    shift = db.relationship('Shift', foreign_keys=[shift_id])
    owner = db.relationship('User', foreign_keys=[owner_id])
    candidate = db.relationship('User', foreign_keys=[candidate_id])

    def to_dict(self):
        s = self.shift
        return {
            'id': self.id,
            'status': self.status,
            'date': s.shift_date.isoformat() if s else None,
            'code': s.shift_code if s else None,
            'owner': self.owner.to_dict() if self.owner else None,
            'candidate': self.candidate.to_dict() if self.candidate else None,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }

# ---------------------------------
# DB init + light migrations
# ---------------------------------
with app.app_context():
    db.create_all()
    from sqlalchemy import inspect
    engine = db.engine
    insp = inspect(engine)
    backend = engine.url.get_backend_name()

    def _exec(sql, ok_msg):
        try:
            db.session.execute(sqltext(sql))
            db.session.commit()
            app.logger.info(ok_msg)
        except Exception as e:
            db.session.rollback()
            app.logger.warning(f"skip: {e}")
        # ensure control_events
    if 'control_events' not in insp.get_table_names():
        _exec("""
            CREATE TABLE IF NOT EXISTS control_events(
              id SERIAL PRIMARY KEY,
              kind VARCHAR(20) NOT NULL,
              user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
              event_date DATE NOT NULL,
              reason TEXT,
              hours NUMERIC(5,2),
              time_from VARCHAR(5),
              time_to VARCHAR(5),
              created_by_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
              created_at TIMESTAMPTZ DEFAULT NOW()
            )""", "control_events ensured")

    # ensure columns on users
    if 'users' in insp.get_table_names():
        ucols = {c['name']: c for c in insp.get_columns('users')}
        if 'hourly_rate_pln' not in ucols:
            _exec("ALTER TABLE users ADD COLUMN hourly_rate_pln NUMERIC", "users.hourly_rate_pln added")
        if 'tax_percent' not in ucols:
            _exec("ALTER TABLE users ADD COLUMN tax_percent NUMERIC DEFAULT 0", "users.tax_percent added")
        if 'order_index' not in ucols:
            _exec("ALTER TABLE users ADD COLUMN order_index INTEGER", "users.order_index added")
        if 'is_coordinator' not in ucols:
            _exec("ALTER TABLE users ADD COLUMN is_coordinator BOOLEAN DEFAULT FALSE NOT NULL", "users.is_coordinator added")
        if 'is_zmiwaka' not in ucols:
            _exec("ALTER TABLE users ADD COLUMN is_zmiwaka BOOLEAN DEFAULT FALSE NOT NULL", "users.is_zmiwaka added")

        # allow null email/password in Postgres
        if backend.startswith('postgresql'):
            cols = {c['name']: c for c in insp.get_columns('users')}
            for col in ('email', 'password_hash'):
                ci = cols.get(col)
                if ci is not None and not ci.get('nullable', True):
                    _exec(f"ALTER TABLE users ALTER COLUMN {col} DROP NOT NULL", f"users.{col} set NULLABLE")

    # dedupe shifts per (user_id, date)
    try:
        dups = (db.session.query(Shift.user_id, Shift.shift_date, func.count(Shift.id))
                .group_by(Shift.user_id, Shift.shift_date)
                .having(func.count(Shift.id) > 1)
                .all())
        for uid, sdate, _ in dups:
            rows = (Shift.query
                    .filter(Shift.user_id == uid, Shift.shift_date == sdate)
                    .order_by(Shift.id.desc())
                    .all())
            for extra in rows[1:]:
                db.session.delete(extra)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f"dedupe shifts skipped: {e}")

    # unique index
    _exec("CREATE UNIQUE INDEX IF NOT EXISTS uq_shifts_user_date ON shifts (user_id, shift_date)",
          "uq_shifts_user_date ensured")

    # ensure swap_proposals table (sqlite path: db.create_all() already did)
    if 'swap_proposals' not in insp.get_table_names() and backend.startswith('postgresql'):
        try:
            db.session.execute(sqltext("""
                CREATE TABLE IF NOT EXISTS swap_proposals (
                    id SERIAL PRIMARY KEY,
                    requester_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    target_user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    my_date DATE NOT NULL,
                    their_date DATE NOT NULL,
                    status VARCHAR(20) NOT NULL DEFAULT 'pending',
                    created_at TIMESTAMPTZ DEFAULT NOW()
                )"""))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.warning(f"swap_proposals migration skipped: {e}")

    # ensure market_offers (for Postgres safety)
    if 'market_offers' not in insp.get_table_names() and backend.startswith('postgresql'):
        try:
            db.session.execute(sqltext("""
                CREATE TABLE IF NOT EXISTS market_offers(
                    id SERIAL PRIMARY KEY,
                    shift_id INTEGER UNIQUE NOT NULL REFERENCES shifts(id) ON DELETE CASCADE,
                    owner_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    candidate_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                    status VARCHAR(20) NOT NULL DEFAULT 'open',
                    created_at TIMESTAMPTZ DEFAULT NOW()
                )
            """))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            app.logger.warning(f"market_offers migration skipped: {e}")
# ensure column 'coord_lounge' exists on shifts
def ensure_coord_lounge_column():
    insp = db.inspect(db.engine)
    cols = [c['name'] for c in insp.get_columns('shifts')]
    if 'coord_lounge' not in cols:
        db.session.execute(text("ALTER TABLE shifts ADD COLUMN coord_lounge VARCHAR(16)"))
        db.session.commit()


def ensure_lounge_column():
    insp = db.inspect(db.engine)
    cols = [c['name'] for c in insp.get_columns('shifts')]
    if 'lounge' not in cols:
        db.session.execute(text("ALTER TABLE shifts ADD COLUMN lounge VARCHAR(16)"))
        db.session.commit()


# ---------------------------------
# Helpers
# ---------------------------------
COORDINATORS_DEFAULT = {
    "Justyna Świstak", "Alicja Palczewska", "Jakub Włodarczyk", "Alicja Kupczyk",
    "Patrycja Gołebiowska", "Roman Wozniak", "Maria Romanova", "Alicja Daniel",
    "Karina Levchenko", "Artem Bilenko", "Wiktoria Utko",
}
ZMIWAKI_DEFAULT = {
    "Tetiana Rudiuk", "Maiia Rybchynchuk", "Marina Prykhodko", "Maryna Prykhodko",
    "Марина Приходько",
}

def current_user():
    try:
        uid = int((get_jwt() or {}).get('sub'))
    except Exception:
        return None
    return db.session.get(User, uid)

def _norm(s: str) -> str:
    if s is None:
        return ''
    s = s.strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s

COORDINATORS_DEFAULT_NORM = {_norm(name) for name in COORDINATORS_DEFAULT}
ZMIWAKI_DEFAULT_NORM = {_norm(name) for name in ZMIWAKI_DEFAULT}

def _is_coord_or_admin(u: User|None) -> bool:
    if not u: return False
    role = (u.role or '').lower()
    return role in ('admin','coordinator') or is_coordinator_user(u)

def is_coordinator_user(user: User | None) -> bool:
    if not user:
        return False
    if bool(getattr(user, 'is_coordinator', False)):
        return True
    role = (getattr(user, 'role', '') or '').lower()
    if role == 'coordinator':
        return True
    return _norm(getattr(user, 'full_name', '')) in COORDINATORS_DEFAULT_NORM


def is_zmiwaka_user(user: User | None) -> bool:
    if not user:
        return False
    if bool(getattr(user, 'is_zmiwaka', False)):
        return True
    return _norm(getattr(user, 'full_name', '')) in ZMIWAKI_DEFAULT_NORM

DATE_PATTERNS = [
    (re.compile(r'^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$'), ('d','m','y')),
    (re.compile(r'^(\d{4})[./-](\d{1,2})[./-](\d{1,2})$'), ('y','m','d')),
]

def code_group(code: str) -> str:
    c = (code or '').strip().upper()
    if c.startswith('1'): return 'morning'
    if c.startswith('2'): return 'evening'
    return ''
# --- helpers (рядом с другими утилитами) ---

def _shift_to_dict(shift, user):
    """Единый формат для фронта. Координатор — если У ЭТОЙ смены есть coord_lounge."""
    code = (shift.shift_code or '').strip()
    lounge = (shift.lounge or '' ).strip().lower() or None
    coord_lounge = (shift.coord_lounge or '').strip().lower() or None

    # bar: если в коде есть 'B' (1/B, 2/B, B и т.п.)
    is_bar = 'B' in code.upper()

    # zmywak: если так помечено (оставь свою проверку, если другая)
    note = (shift.work_note or '').lower()
    is_zmiwak = 'zmyw' in note or 'zmywak' in note

    return {
        'user_id'        : user.id,
        'full_name'      : user.full_name,
        'shift_code'     : code,
        'lounge'         : lounge,        # цвет цифры
        'coord_lounge'   : coord_lounge,  # цвет заливки коорда
        'is_coordinator' : bool(coord_lounge),
        'is_bar_today'   : is_bar,
        'is_zmiwaka'     : is_zmiwak,
    }


# --- Timezone helpers (Europe/Warsaw) ---
try:
    from zoneinfo import ZoneInfo  # py3.9+
    _WARSAW_TZ = ZoneInfo("Europe/Warsaw")
except Exception:
    _WARSAW_TZ = None

def warsaw_today():
    if _WARSAW_TZ:
        return datetime.now(_WARSAW_TZ).date()
    # fallback (чуть менее точно без TZ): системная дата
    return _date.today()

def warsaw_tomorrow():
    return warsaw_today() + timedelta(days=1)
 

def _purge_shifts_and_offers_in_range(start_date, end_date):
    """
    Удаляет market_offers и shifts в указанном диапазоне дат.
    Делает это «по-правильному», чтобы FK не ронял транзакцию.
    """
    # подзапрос айдишников смен по диапазону
    shift_ids_q = db.session.query(Shift.id).filter(
        Shift.shift_date.between(start_date, end_date)
    )

    # сначала — офферы, потом — сами смены
    db.session.execute(
        delete(MarketOffer).where(MarketOffer.shift_id.in_(shift_ids_q))
    )
    db.session.execute(
        delete(Shift).where(Shift.id.in_(shift_ids_q))
    )
    # flush, чтобы гарантированно освободить FK до вставок
    db.session.flush()


def _xlsx_iter_colored_office_cells(xlsx_path: str, year: int, month: int):
    """
    Генератор: (full_name:str, date:date, coord_lounge:'mazurek'|'polonez'|None)
    Берём только клетки со значением '1' или '2' и с заливкой.
    """
    from datetime import date as _date_cls

    def _to_rgb_tuple(color):
        if not color: return None
        rgb = getattr(color, "rgb", None)
        if not rgb: return None
        rgb = str(rgb).upper()
        if len(rgb) == 8:
            r = int(rgb[2:4], 16); g = int(rgb[4:6], 16); b = int(rgb[6:8], 16)
            return (r, g, b)
        if len(rgb) == 6:
            r = int(rgb[0:2], 16); g = int(rgb[2:4], 16); b = int(rgb[4:6], 16)
            return (r, g, b)
        return None

    def _detect_lounge(rgb):
        if not rgb: return None
        r, g, b = rgb
        if r >= 200 and g >= 180 and b <= 140: return "polonez"  # жёлтый
        if b >= 150 and r <= 140 and g <= 160: return "mazurek"  # синий
        return None

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    day_cols = []
    for idx, val in enumerate(header):
        try:
            d = int(str(val).strip())
        except Exception:
            continue
        if 1 <= d <= 31:
            day_cols.append((idx, d))
    if not day_cols:
        return

    for row in ws.iter_rows(min_row=2, values_only=False):
        full_name = (row[0].value or "").strip() if row[0].value else ""
        if not full_name:
            continue
        for idx, day in day_cols:
            cell = row[idx]
            val = (str(cell.value).strip() if cell.value is not None else "")
            if val not in ("1", "2"):
                continue
            rgb = _to_rgb_tuple(cell.fill.start_color)
            lounge = _detect_lounge(rgb)
            if lounge:
                yield full_name, _date_cls(year, month, day), lounge


@app.teardown_appcontext
def teardown_db(exception=None):
    try:
        db.session.remove()
    except Exception:
        pass

def db_get(model, pk):
    return db.session.get(model, pk)

@app.errorhandler(OperationalError)
def on_operational_error(e):
    app.logger.error(f"DB error: {e}")
    return jsonify({'error':'Baza danych przeciążona (połączenia). Spróbuj ponownie za chwilę.'}), 503

@app.errorhandler(DisconnectionError)
def on_disconnection(e):
    app.logger.error(f"DB disconnect: {e}")
    return jsonify({'error':'Baza danych przeciążona (rozłączenie). Spróbuj ponownie.'}), 503

def _parse_hhmm(s: str):
    s = (s or '').strip()
    if not s: return None
    try:
        hh, mm = map(int, s.split(':'))
        if not (0 <= hh < 24 and 0 <= mm < 60): return None
        return hh * 60 + mm
    except Exception:
        return None

def _minutes_to_hours_float(mins: int) -> float:
    return round((mins or 0) / 60.0, 2)

def _has_shift(user_id: int, d: date, exclude_id: int | None = None) -> bool:
    q = Shift.query.filter(Shift.user_id == user_id, Shift.shift_date == d)
    if exclude_id:
        q = q.filter(Shift.id != exclude_id)
    return db.session.query(q.exists()).scalar()

SHIFT_HOURS_DEFAULT = {
    '1': 9.5, '1/B': 9.5, '1B': 9.5,
    '2': 9.5, '2/B': 9.5, '2B': 9.5,
}
def resolve_hours(shift_obj):
    if getattr(shift_obj, 'worked_hours', None) is not None:
        try: return float(shift_obj.worked_hours or 0)
        except: return 0.0
    if shift_obj.hours is not None:
        try: return float(shift_obj.hours or 0)
        except: return 0.0
    code = (shift_obj.shift_code or '').strip().upper().replace(' ', '')
    return float(SHIFT_HOURS_DEFAULT.get(code, 0))

def default_times_for_code(code: str):
    c = (code or '').strip().upper().replace(' ', '')
    if c in ('1','1/B','1B'):
        return '04:30', '14:00'
    if c in ('2','2/B','2B'):
        return '14:00', '23:30'
    return None, None

import re


from werkzeug.utils import secure_filename

@app.post('/api/admin/lounge-from-xlsx')
@jwt_required()
def lounge_from_xlsx():
    # только админ
    me = current_user()  # если у тебя есть util; иначе достань из JWT
    if not me or (me.role or '').lower() != 'admin':
        return jsonify({'error': 'Forbidden'}), 403

    year = int(request.form.get('year', '0') or 0)
    month = int(request.form.get('month', '0') or 0)
    f = request.files.get('file')
    if not f or year < 2000 or year > 2100 or month < 1 or month > 12:
        return jsonify({'error': 'Bad input'}), 400

    tmp_path = os.path.join(tempfile.gettempdir(), secure_filename(f.filename))
    f.save(tmp_path)

    # соберём (full_name,date)->lounge из XLSX
    updates = []
    for full_name, d, lounge in _xlsx_iter_colored_office_cells(tmp_path, year, month):
        updates.append((full_name, d, lounge))

    if not updates:
        return jsonify({'ok': True, 'updated': 0})

    # Применяем к БД: только для координаторов
    updated = 0
    for full_name, d, lounge in updates:
        u = db.session.query(User).filter(User.full_name == full_name).first()
        if not u:
            continue
        if (u.role or '').lower() != 'coordinator':
            continue
        sh = (db.session.query(Shift)
              .filter(Shift.user_id == u.id, Shift.shift_date == d)
              .first())
        if not sh:
            continue
        sh.coord_lounge = lounge  # 'mazurek'|'polonez'
        updated += 1

    db.session.commit()
    try:
        os.remove(tmp_path)
    except Exception:
        pass
    return jsonify({'ok': True, 'updated': updated})



def parse_date_fuzzy(token, default_month=None, default_year=None):
    """
    Принимает '01.08', '1,08', '1/08', '2025-08-01', '01-08-2025', '1' (только день).
    Если пришёл только день — подставит default_month/default_year
    (или текущие из _date.today()).
    Возвращает 'YYYY-MM-DD' либо None.
    """
    if token is None:
        return None
    s = str(token).strip()
    if not s:
        return None

    # нормализуем разделители
    s = s.replace(',', '.').replace('/', '.').replace('-', '.')
    parts = [p for p in re.split(r'\D+', s) if p]

    try:
        if len(parts) == 3:
            # если есть 4-значный год — используем его
            if len(parts[0]) == 4:  # Y M D
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            elif len(parts[2]) == 4:  # D M Y
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            else:  # по умолчанию D M Y
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        elif len(parts) == 2:
            d, m = int(parts[0]), int(parts[1])
            y = int(default_year if default_year is not None else _date.today().year)
        elif len(parts) == 1:
            d = int(parts[0])
            m = int(default_month if default_month is not None else _date.today().month)
            y = int(default_year  if default_year  is not None else _date.today().year)
        else:
            return None

        return _date(y, m, d).isoformat()
    except Exception:
        return None


def _get_user_from_jwt():
    """Возвращает (user_id, User) из JWT, без legacy Query.get()."""
    claims = get_jwt() or {}
    uid = int(claims["sub"])
    u = db_get(User, uid)  # вместо User.query.get(uid)
    return uid, u


def _is_manager(user_or_id) -> bool:
    """Менеджер = роль admin/coordinator ИЛИ email в MANAGER_EMAILS."""
    u = user_or_id
    if isinstance(u, int):
        u = User.query.get(u)
    if not u:
        return False
    email = (u.email or '').strip().lower()
    role  = (u.role  or '').strip().lower()
    return (role in ('admin', 'coordinator')) or (email in MANAGER_EMAILS)

def _shift_of(user_id, d: date):
    return Shift.query.filter(and_(Shift.user_id==user_id, Shift.shift_date==d)).first()


# ---------------------------------
# Auth
# ---------------------------------
@app.post('/api/register')
def register():
    data = request.get_json(force=True)
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    full_name = (data.get('full_name') or '').strip()

    if not email or not password or not full_name:
        return jsonify({'error': 'Pola email, password i full_name są wymagane.'}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Użytkownik z takim adresem email już istnieje.'}), 400

    existing = User.query.filter_by(full_name=full_name).first()
    if existing:
        placeholder = (existing.email or '').endswith('@auto.local')
        if existing.email is None or placeholder:
            existing.email = email
            existing.password_hash = generate_password_hash(password)
            db.session.commit()
            token = create_access_token(identity=str(existing.id),
                                        additional_claims={'role': existing.role, 'full_name': existing.full_name})
            return jsonify({'access_token': token, 'user': existing.to_dict()})
        return jsonify({'error': 'Użytkownik o takim full_name już istnieje.'}), 400

    role = 'admin' if email in ADMIN_EMAILS or User.query.count() == 0 else 'user'
    user = User(email=email, password_hash=generate_password_hash(password), full_name=full_name, role=role)
    db.session.add(user); db.session.commit()

    token = create_access_token(identity=str(user.id),
                                additional_claims={'role': user.role, 'full_name': user.full_name})
    return jsonify({'access_token': token, 'user': user.to_dict()})

@app.post('/api/login')
def login():
    data = request.get_json(force=True)
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    try:
        user = User.query.filter_by(email=email).first()
    except (OperationalError, DisconnectionError):
        return jsonify({'error': 'Baza danych chwilowo niedostępna. Spróbuj ponownie później.'}), 503

    if not user or not user.password_hash or not check_password_hash(user.password_hash, password):
        return jsonify({'error': 'Nieprawidłowy email lub hasło.'}), 401

    token = create_access_token(identity=str(user.id),
                                additional_claims={'role': user.role, 'full_name': user.full_name})
    return jsonify({'access_token': token, 'user': user.to_dict()})

# ---------------------------------
# User settings & stats
# ---------------------------------
@app.get('/api/me/settings')
@jwt_required()
def me_settings_get():
    uid = int(get_jwt()['sub'])
    u = User.query.get(uid)
    return jsonify({
        'hourly_rate_pln': float(u.hourly_rate_pln) if u and u.hourly_rate_pln is not None else None,
        'tax_percent': float(u.tax_percent or 0)
    })

@app.post('/api/me/settings')
@jwt_required()
def me_settings_set():
    uid = int(get_jwt()['sub'])
    u = User.query.get(uid)
    data = request.get_json(force=True)
    try:
        rate = data.get('hourly_rate_pln', None)
        tax  = data.get('tax_percent', 0)
        u.hourly_rate_pln = None if rate in (None, '') else float(rate)
        u.tax_percent     = float(tax or 0)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Błąd zapisu ustawień: {e}'}), 400


@app.post('/api/control/late')
@jwt_required()
def control_add_late():
    me = current_user()
    if not _is_coord_or_admin(me):
        return jsonify({'error':'Forbidden'}), 403
    data = request.get_json(force=True) or {}
    uid  = int(data.get('user_id') or 0)
    date_iso = (data.get('date') or '').strip()
    reason   = (data.get('reason') or '').strip()
    try:
        d = datetime.fromisoformat(date_iso).date()
    except Exception:
        return jsonify({'error':'Bad date'}), 400
    ev = ControlEvent(kind='late', user_id=uid, event_date=d, reason=reason, created_by_id=me.id)
    db.session.add(ev); db.session.commit()
    return jsonify({'ok': True, 'event': ev.to_dict()})
    

@app.post('/api/control/extra')
@jwt_required()
def control_add_extra():
    me = current_user()
    if not _is_coord_or_admin(me):
        return jsonify({'error':'Forbidden'}), 403
    data = request.get_json(force=True) or {}
    uid = int(data.get('user_id') or 0)
    date_iso = (data.get('date') or '').strip()
    reason = (data.get('reason') or '').strip()
    hours  = float(data.get('hours') or 0)
    if hours <= 0:
        return jsonify({'error':'Podaj poprawną liczbę godzin'}), 400
    try:
        d = datetime.fromisoformat(date_iso).date()
    except Exception:
        return jsonify({'error':'Bad date'}), 400
    ev = ControlEvent(kind='extra', user_id=uid, event_date=d, reason=reason, hours=hours, created_by_id=me.id)
    db.session.add(ev); db.session.commit()
    return jsonify({'ok': True, 'event': ev.to_dict()})


@app.post('/api/control/absence')
@jwt_required()
def control_add_absence():
    me = current_user()
    if not _is_coord_or_admin(me):
        return jsonify({'error':'Forbidden'}), 403
    data = request.get_json(force=True) or {}
    uid = int(data.get('user_id') or 0)
    date_iso = (data.get('date') or '').strip()
    reason = (data.get('reason') or '').strip()
    try:
        d = datetime.fromisoformat(date_iso).date()
    except Exception:
        return jsonify({'error':'Bad date'}), 400
    ev = ControlEvent(kind='absence', user_id=uid, event_date=d, reason=reason, created_by_id=me.id)
    db.session.add(ev); db.session.commit()
    return jsonify({'ok': True, 'event': ev.to_dict()})


@app.post('/api/control/add-shift')
@jwt_required()
def control_add_shift():
    me = current_user()
    if not _is_coord_or_admin(me):
        return jsonify({'error':'Forbidden'}), 403
    data = request.get_json(force=True) or {}
    uid = int(data.get('user_id') or 0)
    date_iso = (data.get('date') or '').strip()
    reason = (data.get('reason') or '').strip()
    t_from  = (data.get('from') or '').strip()  # 'HH:MM'
    t_to    = (data.get('to') or '').strip()

    try:
        d = datetime.fromisoformat(date_iso).date()
    except Exception:
        return jsonify({'error':'Bad date'}), 400

    # запретить дубль второй смены в день
    if _has_shift(uid, d):
        return jsonify({'error':'Ten pracownik ma już zmianę w tym dniu.'}), 400

    # вычислим часы
    def _parse_hhmm(s):
        try: h,m = map(int, s.split(':')); return h*60+m
        except: return None
    start = _parse_hhmm(t_from); end = _parse_hhmm(t_to)
    if start is None or end is None:
        return jsonify({'error':'Podaj godziny HH:MM'}), 400
    if end < start: end += 24*60
    hours = round((end-start)/60.0, 2)

    # создаём смену (без кода, но с worked_hours)
    sh = Shift(user_id=uid, shift_date=d, shift_code='X', hours=None, worked_hours=hours, work_note=reason or None)
    db.session.add(sh)

    ev = ControlEvent(kind='manual_shift', user_id=uid, event_date=d, reason=reason,
                      time_from=t_from, time_to=t_to, hours=hours, created_by_id=me.id)
    db.session.add(ev)
    db.session.commit()
    return jsonify({'ok': True, 'event': ev.to_dict(), 'shift_id': sh.id})



@app.get('/api/control/summary')
@jwt_required()
def control_summary():
    # любой залогиненный может смотреть
    month = (request.args.get('month') or '').strip()  # 'YYYY-MM'
    if not month:
        today = _date.today(); y, m = today.year, today.month
    else:
        y, m = map(int, month.split('-'))
    start = _date(y, m, 1)
    end   = _date(y, m, monthrange(y, m)[1])

    # события
    evs = (ControlEvent.query
           .filter(ControlEvent.event_date>=start, ControlEvent.event_date<=end)
           .order_by(ControlEvent.event_date.asc(), ControlEvent.created_at.asc())
           .all())
    events = [e.to_dict() for e in evs]

    # укомплектованность
    # считаем списки людей по каждому дню и слоту
    q = (db.session.query(Shift.shift_date, Shift.shift_code, func.count(Shift.id))
         .filter(Shift.shift_date>=start, Shift.shift_date<=end)
         .group_by(Shift.shift_date, Shift.shift_code))
    counts = {}
    for d, code, cnt in q.all():
        key = d.isoformat()
        slot = 'evening' if str(code or '').strip().upper().startswith('2') else 'morning'
        c = counts.setdefault(key, {'morning':0,'evening':0})
        c[slot] += int(cnt or 0)

    NORM = 12
    staffing = []
    cur = start
    while cur <= end:
        iso = cur.isoformat()
        c = counts.get(iso, {'morning':0,'evening':0})
        staffing.append({
            'date': iso,
            'morning': c['morning'],
            'evening': c['evening'],
            'morning_delta': c['morning'] - NORM,
            'evening_delta': c['evening'] - NORM,
        })
        cur = cur + timedelta(days=1)

    return jsonify({'events': events, 'staffing': staffing})



@app.get('/api/my-stats')
@jwt_required()
def my_stats():
    uid = int(get_jwt()['sub'])
    month = (request.args.get('month') or '').strip()  # 'YYYY-MM'
    if month:
        y, m = map(int, month.split('-'))
        start = date(y, m, 1)
        end   = date(y, m, monthrange(y, m)[1])
    else:
        try:
            start = datetime.fromisoformat(request.args.get('from')).date()
            end   = datetime.fromisoformat(request.args.get('to')).date()
        except Exception:
            return jsonify({'error':'Zły zakres dat.'}), 400

    shifts = (Shift.query
              .filter(Shift.user_id==uid, Shift.shift_date>=start, Shift.shift_date<=end)
              .order_by(Shift.shift_date.asc())
              .all())

    u = User.query.get(uid)
    rate = float(u.hourly_rate_pln) if u and u.hourly_rate_pln is not None else 0.0
    tax  = float(u.tax_percent or 0)

    today = _date.today()
    total_hours = 0
    done_hours  = 0
    daily = []
    for s in shifts:
        h = resolve_hours(s)
        total_hours += h
        is_done = s.shift_date <= today
        if is_done: done_hours += h
        gross = h * rate
        net   = gross * (1 - tax/100.0)
        daily.append({
            'date': s.shift_date.isoformat(),
            'code': s.shift_code,
            'hours': h,
            'done': bool(is_done),
            'gross': round(gross,2),
            'net': round(net,2),
        })

    left_hours = max(total_hours - done_hours, 0)
    gross_done = round(done_hours * rate, 2)
    net_done   = round(gross_done * (1 - tax/100.0), 2)
    gross_all  = round(total_hours * rate, 2)
    net_all    = round(gross_all * (1 - tax/100.0), 2)

    return jsonify({
        'range': {'from': start.isoformat(), 'to': end.isoformat()},
        'rate_pln': rate, 'tax_percent': tax,
        'hours_total': total_hours,
        'hours_done': done_hours,
        'hours_left': left_hours,
        'gross_done': gross_done,
        'net_done': net_done,
        'gross_all': gross_all,
        'net_all': net_all,
        'daily': daily
    })


@app.get('/api/users')
@jwt_required()
def users_list():
    # любой залогиненный может читать список имён
    rows = User.query.order_by(User.full_name.asc()).all()
    return jsonify([{'id': u.id, 'full_name': u.full_name} for u in rows])


# ---------------------------------
# Shifts & notes
# ---------------------------------
@app.get('/api/my-shifts')
@jwt_required()
def my_shifts():
    user_id = int((get_jwt() or {})["sub"])
    shifts = (Shift.query
              .filter(Shift.user_id == user_id)
              .order_by(Shift.shift_date.asc())
              .all())
    return jsonify([s.to_dict() for s in shifts])

@app.get('/api/day-shifts')
@jwt_required()
def day_shifts():
    """Смены за один день: координатор = если у смены есть coord_lounge"""
    d_str = request.args.get('date', '')
    try:
        d = datetime.fromisoformat(d_str).date()
    except Exception:
        return jsonify({'error': 'Nieprawidłowa data.'}), 400

    backend = db.engine.url.get_backend_name()

    base_q = Shift.query.join(User).filter(Shift.shift_date == d)
    if backend.startswith('postgresql'):
        q = base_q.order_by(User.order_index.asc().nulls_last(), User.full_name.asc())
    else:
        from sqlalchemy import case
        q = base_q.order_by(
            case((User.order_index.is_(None), 1), else_=0),
            User.order_index.asc(),
            User.full_name.asc()
        )

    rows = q.all()

    def is_evening(code: str) -> bool:
        return (code or '').strip().upper().startswith('2')

    def is_morning(code: str) -> bool:
        return (code or '').strip().upper().startswith('1')

    morning, evening = [], []
    for s in rows:
        u = s.user
        code = (s.shift_code or '').upper()

        # определяем текущий lounge
        lounge = getattr(s, 'lounge', None)
        coord_lounge = (getattr(s, 'coord_lounge', None) or '').strip().lower() or None

        # логика ролей
        is_bar = 'B' in code
        is_zmiwak = is_zmiwaka_user(u)
        is_coord_today = bool(coord_lounge)  # КЛЮЧЕВОЕ: только если есть coord_lounge

        item = {
            'user_id': u.id if u else None,
            'full_name': (u.full_name if u else None),
            'shift_code': s.shift_code,
            'hours': s.hours,
            'order_index': getattr(u, 'order_index', None),
            'is_coordinator': is_coord_today,
            'is_zmiwaka': is_zmiwak,
            'is_bar_today': is_bar,
            'lounge': lounge,
            'coord_lounge': coord_lounge,
        }

        if is_evening(s.shift_code):
            evening.append(item)
        elif is_morning(s.shift_code):
            morning.append(item)

    return jsonify({'date': d.isoformat(), 'morning': morning, 'evening': evening})



@app.route('/api/admin/control')
@jwt_required()
def admin_control():
    claims = get_jwt()
    if claims.get('role') not in ('admin', 'coordinator'):
        return jsonify({'error': 'Brak uprawnień'}), 403

    ym = request.args.get('month', '')
    year, month = map(int, ym.split('-')) if '-' in ym else (date.today().year, date.today().month)

    # === 1. Swapy ===
    swaps = Proposal.query.filter(
        extract('year', Proposal.created_at) == year,
        extract('month', Proposal.created_at) == month,
        Proposal.status.in_(['accepted', 'approved'])
    ).all()

    swaps_data = [{
        "date": s.created_at.strftime('%Y-%m-%d'),
        "from": s.requester.full_name if s.requester else '?',
        "to": s.target_user.full_name if s.target_user else '?',
        "shift_from": s.my_date,
        "shift_to": s.their_date
    } for s in swaps]

    # === 2. Extra godziny ===
    extra = WorkLog.query.filter(
        extract('year', WorkLog.date) == year,
        extract('month', WorkLog.date) == month,
        WorkLog.worked_hours > WorkLog.default_hours
    ).all()
    extra_data = [{
        "date": w.date.strftime('%Y-%m-%d'),
        "user": w.user.full_name if w.user else '?',
        "shift": w.shift_code,
        "extra": round(w.worked_hours - w.default_hours, 2),
        "note": w.note or ''
    } for w in extra]

    # === 3. Nieobecności ===
    missing = Attendance.query.filter(
        extract('year', Attendance.date) == year,
        extract('month', Attendance.date) == month,
        Attendance.status == 'absent'
    ).all()
    missing_data = [{
        "date": m.date.strftime('%Y-%m-%d'),
        "user": m.user.full_name if m.user else '?',
        "shift": m.shift_code,
        "reason": m.reason or ''
    } for m in missing]

    # === 4. Bilans obsady ===
    shifts = Shift.query.filter(
        extract('year', Shift.shift_date) == year,
        extract('month', Shift.shift_date) == month
    ).all()
    daily = {}
    for s in shifts:
        key = s.shift_date.isoformat()
        if key not in daily:
            daily[key] = {'rano': 0, 'popo': 0}
        if s.shift_code.lower().startswith('1'):
            daily[key]['rano'] += 1
        elif s.shift_code.lower().startswith('2'):
            daily[key]['popo'] += 1
    imbalance_data = []
    for d, v in daily.items():
        if v['rano'] != 12 or v['popo'] != 12:
            imbalance_data.append({
                'date': d,
                'rano': v['rano'],
                'popo': v['popo'],
                'status': 'niedobór' if v['rano'] < 12 or v['popo'] < 12 else 'nadwyżka'
            })

    return jsonify({
        "swaps": swaps_data,
        "extra_hours": extra_data,
        "missing": missing_data,
        "imbalance": imbalance_data
    })
    

@app.get('/api/month-shifts')
@jwt_required()
def month_shifts():
    """Возвращает весь месяц пачкой: {"YYYY-MM-DD": {"morning":[...], "evening":[...]}, ...}.
       Координатор = только если в этой смене есть coord_lounge.
    """
    try:
        y = int(request.args.get('year', '0'))
        m = int(request.args.get('month', '0'))
    except ValueError:
        return jsonify({'error': 'Bad year/month'}), 400

    if y < 2000 or y > 2100 or m < 1 or m > 12:
        return jsonify({'error': 'Bad year/month'}), 400

    first = _date(y, m, 1)
    last  = _date(y, m, monthrange(y, m)[1])

    q = (db.session.query(Shift, User)
         .join(User, User.id == Shift.user_id)
         .filter(Shift.shift_date >= first, Shift.shift_date <= last)
         .order_by(Shift.shift_date.asc(), User.order_index.asc().nulls_last(), User.full_name.asc()))

    out = {}
    for sh, u in q.all():
        iso  = sh.shift_date.isoformat()
        slot = 'evening' if str(sh.shift_code or '').strip().startswith('2') else 'morning'
        out.setdefault(iso, {'morning': [], 'evening': []})

        code = (sh.shift_code or '').upper()
        lounge = getattr(sh, 'lounge', None)
        coord_lounge = (getattr(sh, 'coord_lounge', None) or '').strip().lower() or None

        is_bar = 'B' in code
        is_zmiwak = is_zmiwaka_user(u)
        is_coord_today = bool(coord_lounge)  # <--- фиксация

        out[iso][slot].append({
            'user_id': u.id,
            'full_name': u.full_name,
            'shift_code': sh.shift_code,
            'hours': sh.hours,
            'order_index': getattr(u, 'order_index', None),
            'is_coordinator': is_coord_today,
            'is_zmiwaka': is_zmiwak,
            'is_bar_today': is_bar,
            'lounge': lounge,
            'coord_lounge': coord_lounge,
        })

    return jsonify(out)




@app.get('/api/shifts')
@jwt_required()
def get_all_shifts():
    # Optional ?month=YYYY-MM
    month = (request.args.get('month') or '').strip()
    date_filter = []
    if month:
        y, m = map(int, month.split('-'))
        start = date(y, m, 1)
        end   = date(y, m, monthrange(y, m)[1])
        date_filter = [Shift.shift_date >= start, Shift.shift_date <= end]

    backend = db.engine.url.get_backend_name()
    COORD = set(COORDINATORS_DEFAULT)
    ZMIW  = set(ZMIWAKI_DEFAULT)

    base_q = Shift.query.join(User)
    if date_filter:
        base_q = base_q.filter(and_(*date_filter))

    if backend.startswith('postgresql'):
        q = base_q.order_by(Shift.shift_date.asc(), User.order_index.asc().nulls_last(), User.full_name.asc())
    else:
        from sqlalchemy import case
        q = base_q.order_by(
            Shift.shift_date.asc(),
            case((User.order_index.is_(None), 1), else_=0),
            User.order_index.asc(),
            User.full_name.asc()
        )

    shifts = q.all()
    out = []
    for s in shifts:
        u = s.user
        full_name = u.full_name if u else None
        dct = s.to_dict()
        dct.update({
            'order_index': getattr(u, 'order_index', None),
            'is_coordinator': is_coordinator_user(u),
            'is_zmiwaka': is_zmiwaka_user(u),
        })

        out.append(dct)
    return jsonify(out)

@app.get('/api/my-shifts-brief')
@jwt_required()
def my_shifts_brief():
    uid = int(get_jwt()['sub'])
    month = (request.args.get('month') or '').strip()  # 'YYYY-MM'
    if not month:
        today = _date.today()
        y, m = today.year, today.month
    else:
        y, m = map(int, month.split('-'))
    start = date(y, m, 1)
    end   = date(y, m, monthrange(y, m)[1])

    rows = (Shift.query
            .filter(Shift.user_id==uid, Shift.shift_date>=start, Shift.shift_date<=end)
            .order_by(Shift.shift_date.asc())
            .all())

    out = []
    for s in rows:
        out.append({
            'id': s.id,
            'date': s.shift_date.isoformat(),
            'code': s.shift_code,
            'scheduled_hours': float(s.hours) if s.hours is not None else float(SHIFT_HOURS_DEFAULT.get((s.shift_code or '').strip().upper().replace(' ', ''), 0)),
            'worked_hours': float(s.worked_hours) if s.worked_hours is not None else None,
            'note_preview': (s.work_note[:120] + '…') if s.work_note and len(s.work_note) > 120 else (s.work_note or '')
        })
    return jsonify(out)

@app.get("/api/day/coworkers-shifts")
@jwt_required()
def coworkers_shifts_same_day():
    """
    Возвращает чужие смены в указанный день (для модалки обмена).
    ?date=YYYY-MM-DD
    """
    uid = int(get_jwt()['sub'])
    iso = (request.args.get("date") or "").strip()
    try:
        y, m, d = map(int, iso.split("-"))
        day = date(y, m, d)
    except Exception:
        return jsonify({"error": "bad date"}), 400

    q = (db.session.query(Shift, User)
         .join(User, User.id == Shift.user_id)
         .filter(Shift.shift_date == day, Shift.user_id != uid)
         .order_by(User.full_name.asc()))

    items = []
    for sh, u in q.all():
        items.append({
            "id": sh.id,
            "user_id": u.id,
            "user_name": u.full_name,
            "code": sh.shift_code,
            "date": sh.shift_date.isoformat(),
            "scheduled_hours": float(sh.hours) if sh.hours is not None else float(SHIFT_HOURS_DEFAULT.get((sh.shift_code or '').strip().upper().replace(' ', ''), 0)),
        })
    return jsonify(items)


@app.get('/api/my-shift/<int:sid>')
@jwt_required()
def my_shift_get(sid):
    uid = int(get_jwt()['sub'])
    s = Shift.query.get(sid)
    if not s or s.user_id != uid:
        return jsonify({'error':'Nie znaleziono zmiany.'}), 404
    start_hint, end_hint = default_times_for_code(s.shift_code)
    return jsonify({
        'id': s.id,
        'date': s.shift_date.isoformat(),
        'code': s.shift_code,
        'scheduled_hours': float(s.hours) if s.hours is not None else float(SHIFT_HOURS_DEFAULT.get((s.shift_code or '').strip().upper().replace(' ', ''), 0)),
        'worked_hours': float(s.worked_hours) if s.worked_hours is not None else None,
        'note': s.work_note or '',
        'default_start': start_hint,
        'default_end': end_hint
    })

@app.get('/api/next-shift')
@jwt_required()
def next_shift():
    uid = int(get_jwt()['sub'])
    today = _date.today()
    s = (Shift.query
         .filter(Shift.user_id == uid, Shift.shift_date >= today)
         .order_by(Shift.shift_date.asc())
         .first())
    if not s:
        s = (Shift.query
             .filter(Shift.user_id == uid)
             .order_by(Shift.shift_date.desc())
             .first())
        if not s:
            return jsonify({'empty': True})

    y, m = s.shift_date.year, s.shift_date.month
    month_shifts = (Shift.query
                    .filter(Shift.user_id == uid,
                            Shift.shift_date >= date(y, m, 1),
                            Shift.shift_date <= date(y, m, monthrange(y, m)[1]))
                    .all())
    total = len(month_shifts)
    done  = sum(1 for x in month_shifts if x.shift_date <= today)

    return jsonify({
        'date': s.shift_date.isoformat(),
        'code': s.shift_code,
        'hours': resolve_hours(s),
        'month_total': total,
        'month_done':  done
    })

# -------- Day Notes (single, clean version) --------
from datetime import timezone
from sqlalchemy.sql import func

def _current_uid():
    try:
        return int(get_jwt().get('sub'))
    except Exception:
        return None


class DayNote(db.Model):
    __tablename__ = 'day_notes'
    id = db.Column(db.Integer, primary_key=True)
    note_date = db.Column(db.Date, nullable=False, index=True)
    text = db.Column(db.Text, nullable=False)
    author_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)

    # серверный дефолт — текущее время (UTC) на стороне БД
    created_at = db.Column(db.DateTime(timezone=True), server_default=func.now())

    author = db.relationship('User')

    def to_dict(self):
        # гарантируем tz-aware и отдаём ISO в UTC (Z)
        created = self.created_at
        if created and created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        return {
            'id': self.id,
            'note_date': self.note_date.isoformat(),
            'text': self.text,
            'author_id': self.author_id,
            'author': self.author.full_name if self.author else None,
            'created_at': created.astimezone(timezone.utc).isoformat().replace('+00:00', 'Z') if created else None,
        }



@app.get('/api/day-notes', endpoint='day_notes_list')
@jwt_required()
def day_notes_list():
    d = parse_date_fuzzy(request.args.get('date', ''))
    if not d:
        return jsonify({'error':'Zła data'}), 400
    rows = (DayNote.query
            .filter(DayNote.note_date == d)
            .order_by(DayNote.created_at.asc())
            .all())
    return jsonify([r.to_dict() for r in rows])

@app.post('/api/day-notes', endpoint='day_notes_add')
@jwt_required()
def day_notes_add():
    payload = request.get_json(force=True) or {}
    d = parse_date_fuzzy(payload.get('date'))
    raw = payload.get('text')
    txt = (raw if isinstance(raw, str) else str(raw or '')).strip()
    if not d or not txt:
        return jsonify({'error': 'Brak danych'}), 400

    uid = _current_uid()
    if not uid:
        return jsonify({'error': 'Brak użytkownika'}), 401

    note = DayNote(
        note_date=d,
        text=txt,
        author_id=uid,
        created_at=datetime.now(timezone.utc)   # <<< КЛЮЧЕВОЕ
    )
    db.session.add(note)
    db.session.commit()
    return jsonify(note.to_dict()), 201

@app.delete('/api/day-notes/<int:nid>', endpoint='day_notes_delete')
@jwt_required()
def day_notes_delete(nid):
    # удалять может ТОЛЬКО автор
    uid = _current_uid()
    note = DayNote.query.get(nid)
    if not note:
        return jsonify({'error':'Nie znaleziono'}), 404
    if uid != note.author_id:
        return jsonify({'error':'Tylko autor może usunąć notatkę.'}), 403
    db.session.delete(note); db.session.commit()
    return jsonify({'ok': True})
# -------- /Day Notes --------

# -------- Worklog --------
@app.post('/api/my-shift/<int:sid>/worklog')
@jwt_required()
def my_shift_save_worklog(sid):
    uid = int(get_jwt()['sub'])
    s = Shift.query.get(sid)
    if not s or s.user_id != uid:
        return jsonify({'error':'Nie znaleziono zmiany.'}), 404

    data = request.get_json(force=True)
    worked_hours = data.get('worked_hours', None)
    start_hhmm = data.get('start_time', '')
    end_hhmm   = data.get('end_time', '')
    note       = (data.get('note') or '').strip()

    calc_hours = None
    if (start_hhmm and end_hhmm):
        start_m = _parse_hhmm(start_hhmm)
        end_m   = _parse_hhmm(end_hhmm)
        if start_m is None or end_m is None:
            return jsonify({'error':'Zły format czasu (HH:MM).'}), 400
        if end_m < start_m:
            end_m += 24*60
        calc_hours = _minutes_to_hours_float(end_m - start_m)

    try:
        if worked_hours in (None, ''):
            if calc_hours is not None:
                s.worked_hours = calc_hours
        else:
            s.worked_hours = float(worked_hours)
        s.work_note = note or None
        db.session.commit()
        return jsonify({'ok': True, 'worked_hours': float(s.worked_hours) if s.worked_hours is not None else None})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Błąd zapisu: {e}'}), 400

@app.get('/api/my-notes')
@jwt_required()
def my_notes_month():
    uid = int(get_jwt()['sub'])
    month = (request.args.get('month') or '').strip()  # 'YYYY-MM'
    if not month:
        today = _date.today(); y, m = today.year, today.month
    else:
        y, m = map(int, month.split('-'))
    start = date(y, m, 1); end = date(y, m, monthrange(y, m)[1])

    rows = (Shift.query
            .filter(Shift.user_id==uid, Shift.shift_date>=start, Shift.shift_date<=end, Shift.work_note.isnot(None))
            .order_by(Shift.shift_date.asc())
            .all())
    return jsonify([{'date': s.shift_date.isoformat(), 'note': s.work_note} for s in rows])


# ========== 1. Запрос на сброс ==========
@app.post("/api/password/request")
def password_request():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    user = User.query.filter_by(email=email).first()
    if not user:
        return jsonify({"ok": True})  # не раскрываем, что email не существует

    user.reset_token = secrets.token_urlsafe(32)
    user.reset_expires = datetime.utcnow() + timedelta(minutes=30)
    db.session.commit()

    reset_url = f"https://lot-schedule-api.onrender.com/reset?token={user.reset_token}"
    print("=== RESET LINK ===")
    print(reset_url)
    # Тут можно подключить почту, но пока просто лог в консоль

    return jsonify({"ok": True, "msg": "Reset link sent"})

# ========== 2. Сброс пароля ==========
@app.post("/api/password/reset")
def password_reset():
    data = request.get_json() or {}
    token = data.get("token")
    new_pw = data.get("new_password")

    if not token or not new_pw:
        return jsonify({"error": "Missing token or password"}), 400

    user = User.query.filter_by(reset_token=token).first()
    if not user or not user.reset_expires or user.reset_expires < datetime.utcnow():
        return jsonify({"error": "Invalid or expired token"}), 400

    user.password_hash = bcrypt.generate_password_hash(new_pw).decode("utf-8")
    user.reset_token = None
    user.reset_expires = None
    db.session.commit()
    return jsonify({"ok": True, "msg": "Password updated"})


# ---------------------------------
# Proposals (swap requests) — двухэтапное утверждение
# ---------------------------------
@app.post('/api/proposals')
@jwt_required()
def create_proposal():
    """
    Создать предложение обмена.
    Поддерживает оба формата тела:
      A) { give_shift_id:int, take_shift_id:int }          # РЕКОМЕНДУЕМО (в том числе same-day)
      B) { target_user_id:int, my_date:str, their_date:str }  # старый режим по датам

    Валидация:
      - give принадлежит инициатору; take принадлежит другому пользователю
      - same-day обмен разрешён ТОЛЬКО между различными группами (1↔2)
      - на разных датах проверяем отсутствие второй смены в целевой дате у каждого
      - защита от дубликатных pending
    """
    user_id, _ = _get_user_from_jwt()
    data = request.get_json(force=True) or {}

    # --- Новый путь: по ID смен
    give_id = data.get('give_shift_id')
    take_id = data.get('take_shift_id')
    if give_id and take_id:
        try:
            give = db_get(Shift, int(give_id))
            take = db_get(Shift, int(take_id))
        except Exception:
            give = take = None

        if not give or not take:
            return jsonify({'error': 'shift not found'}), 404
        if give.user_id != user_id:
            return jsonify({'error': 'not your shift'}), 403
        if take.user_id == user_id:
            return jsonify({'error': 'target must be another user'}), 400

        target_user_id = int(take.user_id)
        my_date = give.shift_date
        their_date = take.shift_date

    else:
        # --- Старый путь: по датам
        target_user_id = int(data.get('target_user_id') or 0)
        my_date_iso    = parse_date_fuzzy(data.get('my_date'))
        their_date_iso = parse_date_fuzzy(data.get('their_date'))

        if not target_user_id or not my_date_iso or not their_date_iso:
            return jsonify({'error': 'Pola target_user_id, my_date, their_date są wymagane.'}), 400
        if target_user_id == user_id:
            return jsonify({'error': 'Nie można proponować wymiany samemu sobie.'}), 400

        my_date    = datetime.fromisoformat(my_date_iso).date()
        their_date = datetime.fromisoformat(their_date_iso).date()

        give = _shift_of(user_id, my_date)
        take = _shift_of(target_user_id, their_date)
        if not give:
            return jsonify({'error': 'Nie masz zmiany w tej dacie.'}), 400
        if not take:
            return jsonify({'error': 'Wybrany pracownik nie ma zmiany w tej dacie.'}), 400

    # --- Общая валидация для обоих путей
    tomo = warsaw_tomorrow()
    if my_date < tomo or their_date < tomo:
        return jsonify({'error': 'Wymiany są możliwe tylko od jutra i później.'}), 400

    # same-day: только разные группы (1↔2)
    if my_date == their_date:
        if code_group(give.shift_code) == code_group(take.shift_code):
            return jsonify({'error': 'Wymiana w tym samym dniu możliwa tylko między różnymi zmianami (1↔2).'}), 400
    else:
        # разные дни: у каждого не должно быть второй смены в целевой дате
        if _has_shift(user_id, their_date):
            return jsonify({'error': 'Masz już zmianę w żądanym dniu — wymiana niemożliwa.'}), 400
        if _has_shift(target_user_id, my_date):
            return jsonify({'error': 'Wybrany pracownik ma już zmianę w Twoim dniu — wymiana niemożliwa.'}), 400

    # защита от дублей pending:
    exists = (SwapProposal.query
              .filter_by(requester_id=user_id, target_user_id=target_user_id,
                         my_date=my_date, their_date=their_date, status='pending')
              .first())
    if exists:
        return jsonify({'error': 'Taka propozycja została już wysłana.'}), 400

    sp = SwapProposal(
        requester_id=user_id,
        target_user_id=target_user_id,
        my_date=my_date,
        their_date=their_date,
        status='pending'
    )
    db.session.add(sp); db.session.commit()
    return jsonify({'proposal': sp.to_dict()})


@app.get('/api/proposals')
@jwt_required()
def list_proposals():
    uid, user = _get_user_from_jwt()

    incoming = (SwapProposal.query
                .filter(and_(SwapProposal.target_user_id == uid,
                             SwapProposal.status.in_(['pending', 'accepted', 'declined', 'approved', 'rejected', 'canceled'])))
                .order_by(SwapProposal.created_at.desc())
                .all())

    outgoing = (SwapProposal.query
                .filter(and_(SwapProposal.requester_id == uid,
                             SwapProposal.status.in_(['pending', 'accepted', 'declined', 'approved', 'rejected', 'canceled'])))
                .order_by(SwapProposal.created_at.desc())
                .all())

    resp = {
        'incoming': [p.to_dict() for p in incoming],
        'outgoing': [p.to_dict() for p in outgoing],
    }

    if _is_manager(user):
        queue = (SwapProposal.query
                 .filter(SwapProposal.status == 'accepted')
                 .order_by(SwapProposal.created_at.desc())
                 .all())
        resp['for_approval'] = [p.to_dict() for p in queue]
        resp['to_approve']   = resp['for_approval']  # alias
    else:
        resp['for_approval'] = []

    return jsonify(resp)

@app.post('/api/proposals/<int:pid>/cancel')
@jwt_required()
def cancel_proposal(pid):
    user_id, _ = _get_user_from_jwt()
    p = SwapProposal.query.get(pid)
    if not p: return jsonify({'error': 'Nie znaleziono.'}), 404
    if p.requester_id != user_id: return jsonify({'error': 'Tylko autor może anulować.'}), 403
    if p.status != 'pending': return jsonify({'error': 'Propozycja została już rozpatrzona.'}), 400
    p.status = 'canceled'
    db.session.commit()
    return jsonify({'proposal': p.to_dict()})

@app.post('/api/proposals/<int:pid>/decline')
@jwt_required()
def decline_proposal(pid):
    user_id, _ = _get_user_from_jwt()
    p = SwapProposal.query.get(pid)
    if not p: return jsonify({'error': 'Nie znaleziono.'}), 404
    if p.target_user_id != user_id: return jsonify({'error': 'Możesz odrzucać tylko swoje przychodzące propozycje.'}), 403
    if p.status != 'pending': return jsonify({'error': 'Propozycja została już rozpatrzona.'}), 400
    p.status = 'declined'
    db.session.commit()
    return jsonify({'proposal': p.to_dict()})

@app.post('/api/proposals/<int:pid>/accept')
@jwt_required()
def accept_proposal(pid):
    # Получатель подтверждает → только статус accepted, БЕЗ обмена сменами
    user_id, _ = _get_user_from_jwt()
    p = SwapProposal.query.get(pid)
    if not p: 
        return jsonify({'error': 'Nie znaleziono.'}), 404
    if p.target_user_id != user_id: 
        return jsonify({'error': 'Możesz akceptować tylko swoje przychodzące propozycje.'}), 403
    if p.status != 'pending': 
        return jsonify({'error': 'Propozycja została już rozpatrzona.'}), 400

    s_req = _shift_of(p.requester_id, p.my_date)
    s_tgt = _shift_of(p.target_user_id, p.their_date)
    if not s_req or not s_tgt:
        return jsonify({'error': 'Zmiany uległy zmianie — wymiana niemożliwa.'}), 409

    if p.my_date == p.their_date:
        # тот же день — разрешено только 1↔2
        if code_group(s_req.shift_code) == code_group(s_tgt.shift_code):
            return jsonify({'error': 'W tym samym dniu wymiana tylko między 1 i 2.'}), 409
    else:
        # разные дни — защита от «третьих» смен
        if _has_shift(p.requester_id, p.their_date):
            return jsonify({'error': 'Autor ma już zmianę w docelowym dniu.'}), 409
        if _has_shift(p.target_user_id, p.my_date):
            return jsonify({'error': 'Pracownik ma już zmianę w docelowym dniu.'}), 409

    p.status = 'accepted'
    db.session.commit()
    return jsonify({'proposal': p.to_dict()})

@app.post('/api/proposals/<int:pid>/approve')
@jwt_required()
def approve_proposal(pid):
    _, user = _get_user_from_jwt()
    if not _is_manager(user):
        return jsonify({'error': 'Tylko przełożony może zatwierdzać.'}), 403

    p = db_get(SwapProposal, pid)
    if not p:
        return jsonify({'error': 'Nie znaleziono.'}), 404
    if p.status != 'accepted':
        return jsonify({'error': 'Propozycja nie jest w stanie do zatwierdzenia.'}), 400

    s_req = _shift_of(p.requester_id, p.my_date)
    s_tgt = _shift_of(p.target_user_id, p.their_date)

    def reject_with(reason, code=409):
        try:
            p.status = 'rejected'
            db.session.commit()
        except Exception:
            db.session.rollback()
        return jsonify({'error': reason, 'proposal': p.to_dict()}), code

    if not s_req or not s_tgt:
        return reject_with('Zmiany uległy zmianie — wymiana niemożliwa.')

    same_day = (p.my_date == p.their_date)

    # same-day допускаем только если разные группы (1↔2)
    if same_day and code_group(s_req.shift_code) == code_group(s_tgt.shift_code):
        return reject_with('W tym samym dniu wymiana tylko między różnymi zmianami (1↔2).')

    if same_day:
        # --- SAME DAY: меняем коды и часы, user_id не трогаем
        s_req.shift_code, s_tgt.shift_code = s_tgt.shift_code, s_req.shift_code
        s_req.hours,      s_tgt.hours      = s_tgt.hours,      s_req.hours
        p.status = 'approved'
        db.session.commit()
        return jsonify({'proposal': p.to_dict()})

    # --- РАЗНЫЕ ДАТЫ: обычный своп владельцев
    # корректная проверка дублей (без учёта самих текущих записей)
    if _has_shift(p.requester_id, p.their_date):
        return reject_with('Autor ma już zmianę w docelowym dniu.')
    if _has_shift(p.target_user_id, p.my_date):
        return reject_with('Pracownik ma już zmianę w docelowym dniu.')

    try:
        s_req.user_id, s_tgt.user_id = p.target_user_id, p.requester_id
        p.status = 'approved'
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return reject_with('Zmiany uległy zmianie — wymiana niemożliwa.', code=409)

    return jsonify({'proposal': p.to_dict()})

@app.post('/api/proposals/<int:pid>/reject')
@jwt_required()
def reject_proposal(pid):
    _, user = _get_user_from_jwt()
    if not _is_manager(user):
        return jsonify({'error': 'Tylko przełożony może odrzucać.'}), 403
    p = SwapProposal.query.get(pid)
    if not p:
        return jsonify({'error': 'Nie znaleziono.'}), 404
    if p.status != 'accepted':
        return jsonify({'error': 'Propozycja nie jest w stanie do odrzucenia.'}), 400
    p.status = 'rejected'
    db.session.commit()
    return jsonify({'proposal': p.to_dict()})

# ---------------------------------
# Market (Rynek zmian) + Takeovers
# ---------------------------------
@app.post('/api/market/offers/<int:shift_id>')
@jwt_required()
def market_offer_create(shift_id):
    uid = int(get_jwt()['sub'])
    sh = Shift.query.get(shift_id)
    if not sh:
        return jsonify({'error':'Nie znaleziono zmiany'}), 404
    if sh.user_id != uid and (get_jwt().get('role','').lower() != 'admin'):
        return jsonify({'error':'To nie jest Twoja zmiana'}), 403

    # уже есть оффер? — не роняем 500
    existing = MarketOffer.query.filter_by(shift_id=shift_id).first()
    if existing:
        return jsonify({'error': 'Ta zmiana już jest na rynku', 'offer_id': existing.id}), 409

    mo = MarketOffer(shift_id=shift_id, owner_id=uid, candidate_id=None, status='open')
    db.session.add(mo)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        existing = MarketOffer.query.filter_by(shift_id=shift_id).first()
        if existing:
            return jsonify({'error': 'Ta zmiana już jest na rynku', 'offer_id': existing.id}), 409
        raise

    return jsonify({'ok': True, 'offer_id': mo.id})


@app.get('/api/market/offers')
@jwt_required()
def market_offers_list():
    uid = int(get_jwt()['sub'])
    open_offers = (MarketOffer.query
                   .filter(MarketOffer.status=='open', MarketOffer.owner_id != uid)
                   .order_by(MarketOffer.created_at.desc())
                   .all())
    my_offers = (MarketOffer.query
                 .filter(MarketOffer.owner_id==uid, MarketOffer.status.in_(['open','requested','approved','rejected']))
                 .order_by(MarketOffer.created_at.desc())
                 .all())
    return jsonify({
        'open': [o.to_dict() for o in open_offers],
        'mine': [o.to_dict() for o in my_offers]
    })

@app.post('/api/market/offers/<int:oid>/claim')
@jwt_required()
def market_offer_claim(oid):
    uid = int(get_jwt()['sub'])
    o = MarketOffer.query.get(oid)
    if not o or o.status != 'open':
        return jsonify({'error':'Oferta niedostępna.'}), 404
    if o.owner_id == uid:
        return jsonify({'error':'To jest Twoja oferta.'}), 400
    # У кандидата не должно быть смены в этот день
    day = o.shift.shift_date
    if _has_shift(uid, day):
        return jsonify({'error':'Masz już zmianę w tym dniu.'}), 400
    shift = o.shift
    if shift and shift.shift_date < warsaw_tomorrow():
        return jsonify({'error': 'Nie można wziąć zmiany z przeszłości ani z dzisiaj.'}), 400

    o.candidate_id = uid
    o.status = 'requested'
    db.session.commit()
    return jsonify({'offer': o.to_dict()})

@app.post('/api/market/offers/<int:oid>/cancel')
@jwt_required()
def market_offer_cancel(oid):
    uid = int(get_jwt()['sub'])
    o = MarketOffer.query.get(oid)
    if not o or o.owner_id != uid:
        return jsonify({'error':'Nie znaleziono lub brak uprawnień.'}), 404
    if o.status != 'open':
        return jsonify({'error':'Tę ofertę nie można anulować.'}), 400
    o.status = 'cancelled'
    db.session.commit()
    return jsonify({'offer': o.to_dict()})

@app.post('/api/market/offers/<int:oid>/approve')
@jwt_required()
def market_offer_approve(oid):
    uid = int(get_jwt()['sub'])
    o = MarketOffer.query.get(oid)
    if not o or o.owner_id != uid:
        return jsonify({'error':'Nie znaleziono lub brak uprawnień.'}), 404
    if o.status != 'requested' or not o.candidate_id:
        return jsonify({'error':'Brak kandydata do zatwierdzenia.'}), 400
    # кандидат всё ещё свободен в этот день?
    day = o.shift.shift_date
    if _has_shift(o.candidate_id, day):
        o.status = 'rejected'
        db.session.commit()
        return jsonify({'error':'Kandydat ma już zmianę w tym dniu.', 'offer': o.to_dict()}), 409
    # перенос смены
    o.shift.user_id = o.candidate_id
    o.status = 'approved'
    db.session.commit()
    return jsonify({'offer': o.to_dict()})

@app.post('/api/market/offers/<int:oid>/reject')
@jwt_required()
def market_offer_reject(oid):
    uid = int(get_jwt()['sub'])
    o = MarketOffer.query.get(oid)
    if not o or o.owner_id != uid:
        return jsonify({'error':'Nie znaleziono lub brak uprawnień.'}), 404
    if o.status != 'requested':
        return jsonify({'error':'Oferta nie oczekuje na kandydata.'}), 400
    o.status = 'open'
    o.candidate_id = None
    db.session.commit()
    return jsonify({'offer': o.to_dict()})

@app.post('/api/takeovers')
@jwt_required()
def takeovers_request():
    """
    Взять чужую смену:
    - если у кандидата уже есть смена в этот день → 400
    - создаём/переиспользуем MarketOffer на shift (owner = владелец смены)
      и ставим candidate_id = requester, status='requested'
    - затем владелец может zatwierdzić/odrzucić через /api/market/offers/<id>/approve|reject
    """
    uid = int(get_jwt()['sub'])
    user = User.query.get(uid)

    data = request.get_json(force=True)
    target_user_id = int(data.get('target_user_id') or 0)
    day = parse_date_fuzzy(data.get('date'))
    if not target_user_id or not day:
        return jsonify({'error':'Brak danych'}), 400

    shift = Shift.query.filter_by(user_id=target_user_id, shift_date=day).first()
    if not shift:
        return jsonify({'error':'Nie znaleziono zmiany.'}), 404

    # у кандидата не должно быть смены в этот день
    if _has_shift(uid, day):
        return jsonify({'error':'Masz już zmianę w tym dniu.'}), 400

    offer = MarketOffer.query.filter_by(shift_id=shift.id).first()
    if not offer:
        offer = MarketOffer(shift_id=shift.id, owner_id=target_user_id,
                            candidate_id=uid, status='requested')
        db.session.add(offer)
    else:
        if offer.status == 'approved':
            return jsonify({'error': 'Ta zmiana została już przekazana.'}), 400
        if offer.status == 'requested' and offer.candidate_id and offer.candidate_id != uid:
            return jsonify({'error': 'Ktoś już poprosił o tę zmianę.'}), 409
        offer.owner_id = target_user_id
        offer.candidate_id = uid
        offer.status = 'requested'

    # уведомим владельца (если есть хелпер)
    try:
        notify_market_request(target_user_id,
                              user.full_name if user else 'Użytkownik',
                              day.isoformat(),
                              shift.shift_code)
    except Exception:
        pass

    db.session.commit()
    return jsonify({'offer': offer.to_dict()})

# ---------------------------------
# Imports (PDF / text)
# ---------------------------------
def _parse_table_no_dates(text: str, year: int, month: int):
    """
    Разбирает «строчный» график БЕЗ строки дат.
    Берёт последние N токенов каждой строки (N = число дней в месяце).
    Возвращает список словарей: {name, day, code}.
    Игнорит строки PLAN/BRAKI/Nazwisko...
    """
    days = monthrange(year, month)[1]
    rows_out = []

    ignore_prefixes = (
        'plan', 'braki', 'nazwisko', 'nazwiska', 'nazwisko i imię', 'nazwisko i imie'
    )

    for raw in text.splitlines():
        ln = raw.strip()
        if not ln:
            continue
        low = ln.lower()
        if any(low.startswith(pfx) for pfx in ignore_prefixes):
            continue

        parts = ln.split()
        if not parts:
            continue

        # если в конце стоит суммарное число (>2), убираем его
        if re.fullmatch(r'\d+', parts[-1]) and int(parts[-1]) > 2:
            parts = parts[:-1]

        if len(parts) <= days:
            # слишком мало столбцов – пропускаем
            continue

        tokens = parts[-days:]                    # последние N токенов -> на дни 1..N
        name   = ' '.join(parts[:-days]).strip()  # всё, что слева – это имя

        if not name:
            # бывает криво распарсили – тогда пропускаем
            continue

        # пробегаем по дням, создаём записи только там, где не 'x'
        for day_idx, code in enumerate(tokens, start=1):
            c = code.strip()
            if not c or c.lower() in ('x', '—', '-', 'x.', 'x,'):
                continue
            rows_out.append({'name': name, 'day': day_idx, 'code': c})

    return rows_out


def map_headers(header_row):
    idx = {'name': None, 'date': None, 'shift': None, 'hours': None}
    NAME_KEYS = {"name","full name","fullname","full_name","employee","pracownik","nazwisko i imię","imię"}
    DATE_KEYS = {"date","data","termin","dzień","dzien"}
    SHIFT_KEYS = {"shift","zmiana","shift_code","kod"}
    HOURS_KEYS = {"hours","godziny"}
    for i, h in enumerate(header_row):
        key = _norm(str(h))
        if key in NAME_KEYS and idx['name'] is None:
            idx['name'] = i
        elif key in DATE_KEYS and idx['date'] is None:
            idx['date'] = i
        elif key in SHIFT_KEYS and idx['shift'] is None:
            idx['shift'] = i
        elif key in HOURS_KEYS and idx['hours'] is None:
            idx['hours'] = i
    ok = idx['name'] is not None and idx['date'] is not None
    return idx if ok else None

def extract_rows_from_pdf(file_bytes: bytes):
    rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                header = [(c or '').strip() for c in table[0]]
                mapping = map_headers(header)
                if not mapping:
                    continue
                for raw in table[1:]:
                    raw = list(raw) + [None] * (len(header) - len(raw))
                    name = (raw[mapping['name']] or '').strip()
                    dstr = (raw[mapping['date']] or '').strip()
                    shift = (raw[mapping['shift']] or '').strip() if mapping['shift'] is not None else ''
                    hours_val = None
                    if mapping['hours'] is not None and raw[mapping['hours']]:
                        try:
                            hours_val = int(str(raw[mapping['hours']]).strip())
                        except Exception:
                            hours_val = None
                    rows.append({'name': name, 'date': dstr, 'shift': shift, 'hours': hours_val})
    return rows

def extract_rows_from_pasted_text(text: str, month_hint: int | None = None):
    """
    Возвращает список словарей:
      {'name': 'Imię Nazwisko', 'date': 'YYYY-MM-DD', 'shift': '1'|'2'|'1/B'|'2/B', 'hours': 9.5}
    Если в тексте есть шапка с датами (dd,MM|dd.MM|dd/MM) — используем её.
    Если шапки нет, но передан month_hint (1..12) — парсим коды по дням 1..N.
    """
    text = (text or '').replace('\t',' ')
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    # выбросим заголовок "Nazwisko i imię"
    lines = [ln for ln in lines if not re.match(r'(?i)^\s*nazwisko.*imi[eę]', ln)]

    date_pat_full = re.compile(r'\b(\d{1,2})[.,/](\d{1,2})\b')
    header_idx = None
    header_tokens = None
    for i, ln in enumerate(lines[:10]):
      toks = date_pat_full.findall(ln)
      if len(toks) >= 8:
        header_idx = i
        header_tokens = [(int(d), int(m)) for d,m in toks]
        break

    year = _date.today().year
    dates: list[datetime.date] = []

    if header_tokens:
        # определим «нормальный» месяц по большинству
        month = Counter([m for _,m in header_tokens]).most_common(1)[0][0]
        # составим список дат текущего года по этому месяцу
        days = [d for d,m in header_tokens if m == month and 1 <= d <= 31]
        dates = [datetime.date(year, month, d) for d in days]
        # вырежем строку шапки
        del lines[header_idx]
    elif month_hint:
        # жёсткий месяц: будем маппить токены по порядку на дни 1..31
        month = int(month_hint)
        if month < 1 or month > 12:
            raise ValueError("Niepoprawny miesiąc.")
        dates = [datetime.date(year, month, d) for d in range(1, 32)]
    else:
        raise ValueError("Nie wykryto poprawnych dat w tekście (brakuje wiersza z datami).")

    # коды смен
    code_re = re.compile(r'(?:1\/B|2\/B|1|2|x|X)')

    rows: list[dict] = []
    for ln in lines:
        if re.match(r'(?i)^(plan|braki)\b', ln):
            continue
        m = code_re.search(ln)
        if not m:
            continue
        name = ln[:m.start()].strip()
        if not name:
            continue
        codes = code_re.findall(ln[m.start():])

        if header_tokens:
            # ровно по количеству дат из шапки
            codes = (codes + ['x'] * len(dates))[:len(dates)]
            date_iter = dates
        else:
            # месяц задан вручную: первые до 31 кодов — это дни 1..31
            codes = codes[:31]
            date_iter = dates[:len(codes)]

        for d, code in zip(date_iter, codes):
            if code.lower() == 'x':
                continue
            if code not in ('1','2','1/B','2/B'):
                continue
            rows.append({
                'name': name,
                'date': d.isoformat(),
                'shift': code,
                'hours': 9.5
            })

    return rows

# ====== ADVANCED PDF PARSER (цвет цифры + цвет заливки) ======
from typing import Tuple, Optional

def _norm_rgb(c):
    """
    pdfplumber возвращает цвет как:
      - None
      - число серого [0..1]
      - кортеж (r,g,b) в [0..1]
    Преобразуем в (R,G,B) 0..255
    """
    if c is None:
        return None
    if isinstance(c, (int, float)):
        v = max(0, min(1, float(c)))
        v = int(round(v * 255))
        return (v, v, v)
    try:
        r, g, b = c[:3]
        r = int(round(max(0, min(1, float(r))) * 255))
        g = int(round(max(0, min(1, float(g))) * 255))
        b = int(round(max(0, min(1, float(b))) * 255))
        return (r, g, b)
    except Exception:
        return None

def _is_blue(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    # голубой/синий: доминирует B, достаточно яркий
    return (b - max(r,g) >= 40) and b >= 140

def _is_black(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    return (r+g+b) <= 120  # тёмный

def _is_yellow(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    # жёлтый: высокие R и G, низкий B
    return (r >= 200 and g >= 180 and b <= 140)

def _bbox_overlaps(a, b) -> bool:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    return not (ax1 <= bx0 or bx1 <= ax0 or ay1 <= by0 or by1 <= ay0)

def _cell_chars(chars, bbox):
    x0,y0,x1,y1 = bbox
    out = []
    for ch in chars:
        cx0, cy0, cx1, cy1 = ch.get('x0'), ch.get('top'), ch.get('x1'), ch.get('bottom')
        if cx0 is None: 
            continue
        # небольшой допуск
        if (cx1 > x0-0.5 and cx0 < x1+0.5 and cy1 > y0-0.5 and cy0 < y1+0.5):
            out.append(ch)
    return out

def _majority_color(chars):
    # берём цвет цифр/символов, усреднённо
    rgbs = []
    for ch in chars:
        rgb = _norm_rgb(ch.get('non_stroking_color'))
        if rgb: rgbs.append(rgb)
    if not rgbs:
        return None
    # среднее
    r = sum(c[0] for c in rgbs)//len(rgbs)
    g = sum(c[1] for c in rgbs)//len(rgbs)
    b = sum(c[2] for c in rgbs)//len(rgbs)
    return (r,g,b)

def _cell_fill_color(rects, bbox):
    # ищем прямоугольник заливки, перекрывающий ячейку
    for r in rects:
        rb = (r.get('x0'), r.get('top'), r.get('x1'), r.get('bottom'))
        if None in rb: 
            continue
        if _bbox_overlaps(bbox, rb):
            rgb = _norm_rgb(r.get('non_stroking_color'))
            if rgb:
                return rgb
    return None

def _parse_schedule_pdf_advanced(file_bytes: bytes, year: int, month: int):
    """
    Возвращает записи:
      {'name', 'date'(YYYY-MM-DD), 'shift', 'lounge', 'coord_lounge'}
    lounge: 'mazurek'|'polonez'|None
    coord_lounge: 'mazurek'|'polonez'|None
    """
    import pdfplumber
    from datetime import date as _date_cls

    out = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            chars = page.chars or []
            rects = page.rects or []

            # 1) найдём заголовок с датами (1..31) — возьмём их X-координаты как границы колонок
            # Берём первую строку, где >= 10 токенов вида числа 1..31
            date_words = [w for w in words if re.fullmatch(r'\d{1,2}', w.get('text','').strip())]
            # сгруппируем по y (строки)
            rows_by_y = {}
            for w in date_words:
                key = round(w['top'] / 5)  # грубая кластеризация по высоте
                rows_by_y.setdefault(key, []).append(w)
            header_row = None
            for _, arr in sorted(rows_by_y.items(), key=lambda kv: len(kv[1]), reverse=True):
                vals = [int(w['text']) for w in arr if 1 <= int(w['text']) <= 31]
                if len(vals) >= 10:  # достаточно столбцов
                    header_row = sorted(arr, key=lambda w: w['x0'])
                    break
            if not header_row:
                continue

            # Список границ X для каждого дня (берём центр слова)
            x_centers = [ (w['x0'] + w['x1'])/2.0 for w in header_row ]
            days      = [ int(w['text']) for w in header_row ]
            # отсортировать на всякий случай
            cols = sorted(zip(days, x_centers), key=lambda t: t[0])
            # фильтр по нашему месяцу (1..N)
            cols = [(d,x) for (d,x) in cols if 1 <= d <= 31]
            if not cols:
                continue

            # превратим точки в интервалы (границы колонок)
            # граница для дня d — середина между центрами d-1 и d
            xs = [x for _,x in cols]
            xbounds = []
            for i,x in enumerate(xs):
                left  = xs[i-1] + (x - xs[i-1]) * 0.5 if i>0 else x - 8   # небольшой fallback
                right = xs[i+1] - (xs[i+1] - x) * 0.5 if i<len(xs)-1 else x + 8
                xbounds.append((left, right))

            # 2) найдём строки сотрудников: берём слова слева от первой колонки и тянем по Y
            first_col_left = min(l for l,_ in xbounds) - 5
            name_candidates = [w for w in words if w['x1'] <= first_col_left and len(w['text'].strip())>1]
            # сгруппируем по строкам (y)
            name_rows = {}
            for w in name_candidates:
                key = round(w['top'] / 2)  # плотнее, имена часто в 2 слова
                name_rows.setdefault(key, []).append(w)

            # для каждой строки — имя и y-границы строки
            rows = []
            for _, arr in name_rows.items():
                arr = sorted(arr, key=lambda w: w['x0'])
                name = ' '.join(w['text'] for w in arr).strip()
                if not name or re.match(r'(?i)^(plan|braki|nazwisko)', name):
                    continue
                top = min(w['top'] for w in arr)
                bottom = max(w['bottom'] for w in arr)
                rows.append((name, top-1, bottom+1))
            # отсортируем сверху-вниз
            rows.sort(key=lambda r: r[1])

            # 3) обходим ячейки: (row, day)
            for name, y0, y1 in rows:
                for (day,(xl,xr)), (_,xc) in zip(enumerate(xbounds, start=1), cols):
                    try:
                        d = _date_cls(year, month, day)
                    except Exception:
                        continue
                    bbox = (xl, y0, xr, y1)

                    cell_ch = _cell_chars(chars, bbox)
                    # вытащим код смены по видимому тексту
                    txt = ''.join(ch['text'] for ch in sorted(cell_ch, key=lambda c: (c['x0'], c['top']))).strip()
                    if not txt:
                        continue
                    # нормализуем: допустим варианты '1', '2', '1/B', '2/B', '1 B'
                    txt_norm = txt.upper().replace(' ', '')
                    if txt_norm not in ('1','2','1/B','2/B','1B','2B'):
                        # игнорируем мусор и 'X'
                        continue
                    shift_code = '1/B' if txt_norm in ('1/B','1B') else ('2/B' if txt_norm in ('2/B','2B') else txt_norm)

                    # цвет цифр → lounge
                    rgb_text = _majority_color(cell_ch)
                    lounge = 'mazurek' if _is_blue(rgb_text) else ('polonez' if _is_black(rgb_text) else None)

                    # заливка ячейки → координатор
                    fill_rgb = _cell_fill_color(rects, bbox)
                    coord_lounge = None
                    if _is_blue(fill_rgb):
                        coord_lounge = 'mazurek'
                    elif _is_yellow(fill_rgb):
                        coord_lounge = 'polonez'

                    out.append({
                        'name': name,
                        'date': d.isoformat(),
                        'shift': shift_code,
                        'lounge': lounge,
                        'coord_lounge': coord_lounge
                    })
    return out

@app.post('/api/upload-pdf-adv')
@jwt_required()
def upload_pdf_advanced():
    """Импорт PDF со считыванием цвета цифры (локация) и заливки (координатор)."""
    claims = get_jwt() or {}
    if (claims.get('role') or '').lower() != 'admin':
        return jsonify({'error': 'Tylko administrator może przesyłać PDF.'}), 403

    f = request.files.get('file')
    year  = int(request.form.get('year') or 0)
    month = int(request.form.get('month') or 0)
    if not f or year < 2000 or not (1 <= month <= 12):
        return jsonify({'error': 'Podaj plik, rok i miesiąc.'}), 400

    data = f.read()
    try:
        rows = _parse_schedule_pdf_advanced(data, year, month)
    except Exception as e:
        return jsonify({'error': f'Błąd odczytu PDF: {e}'}), 400

    # Сбросим только нужный месяц (и офферы)
    first = _date(year, month, 1)
    last  = _date(year, month, monthrange(year, month)[1])

    shift_ids = [s.id for s in Shift.query.filter(Shift.shift_date>=first, Shift.shift_date<=last).all()]
    if shift_ids:
        try:
            MarketOffer.query.filter(MarketOffer.shift_id.in_(shift_ids)).delete(synchronize_session=False)
        except Exception:
            db.session.execute(db.text("DELETE FROM market_offers WHERE shift_id = ANY(:ids)"), {'ids': shift_ids})
        Shift.query.filter(Shift.id.in_(shift_ids)).delete(synchronize_session=False)
    db.session.commit()

    users_by_key = {_norm(u.full_name): u for u in User.query.all()}
    imported = 0
    created_users = []

    for r in rows:
        name = r['name'].strip()
        if not name: 
            continue
        key = _norm(name)
        u = users_by_key.get(key)
        if not u:
            u = User(full_name=name, role='user')
            db.session.add(u); db.session.flush()
            users_by_key[key] = u
            created_users.append(u.full_name)

        sh = Shift(
            user_id = u.id,
            shift_date = r['date'],
            shift_code = r['shift'],
            hours = None,
        )
        # Если в ячейке отмечен координатор — ставим на смену
        if r.get('coord_lounge'):
            sh.coord_lounge = r['coord_lounge']
        db.session.add(sh)
        imported += 1

    db.session.commit()
    return jsonify({'imported': imported, 'created_users': created_users})

# ===== XLSX schedule import (with colors) =====
from typing import Optional, Tuple
from openpyxl import load_workbook

def _rgb_from_openpyxl(color) -> Optional[Tuple[int,int,int]]:
    """
    openpyxl color can be theme, indexed, or ARGB like 'FFRRGGBB'/'RRGGBB'.
    Возвращает (R,G,B) в 0..255 или None.
    """
    if not color:
        return None
    # try .rgb first
    rgb = getattr(color, "rgb", None)
    if rgb:
        rgb = str(rgb).upper()
        if len(rgb) == 8:  # AARRGGBB
            r = int(rgb[2:4], 16); g = int(rgb[4:6], 16); b = int(rgb[6:8], 16)
            return (r,g,b)
        if len(rgb) == 6:  # RRGGBB
            r = int(rgb[0:2], 16); g = int(rgb[2:4], 16); b = int(rgb[4:6], 16)
            return (r,g,b)
    # sometimes theme/indexed -> no direct rgb
    return None

def _is_blue(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    return b >= 150 and (b - max(r,g)) >= 30

def _is_black(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    return (r+g+b) <= 120  # тёмное/чёрное

def _is_yellow(rgb: Optional[Tuple[int,int,int]]) -> bool:
    if not rgb: return False
    r,g,b = rgb
    return r >= 200 and g >= 180 and b <= 140

def _normalize_code(raw: str) -> Optional[str]:
    """
    '1', '2', '1/B', '2/B', '1B', '2B', '1 B', 'X' -> нормализует.
    Возвращает None если пусто/выходной.
    """
    if raw is None: return None
    s = str(raw).strip().upper().replace(' ', '')
    if not s or s in {'X','—','-'}:
        return None
    if s in {'1','2'}: return s
    if s in {'1/B','1B'}: return '1/B'
    if s in {'2/B','2B'}: return '2/B'
    return None

def parse_schedule_xlsx(xlsx_bytes: bytes, year: int, month: int):
    """
    Возвращает list[dict]: {name, date:'YYYY-MM-DD', shift:'1|2|1/B|2/B', lounge, coord_lounge}
    lounge: голубой цвет цифры -> 'mazurek', чёрный -> 'polonez'
    coord_lounge: голубая заливка -> 'mazurek', жёлтая -> 'polonez'
    """
    from datetime import date as _date_cls
    from calendar import monthrange

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # --- 1) найдём строку с заголовками дней (1..31) ---
    header_row = None
    day_cols: list[tuple[int,int]] = []  # [(col_idx_1based, day)]
    # пробуем первые 5 строк
    for r in range(1, 6):
        row_vals = [c.value for c in ws[r]]
        tmp = []
        for j, val in enumerate(row_vals, start=1):  # 1-based
            try:
                d = int(str(val).strip())
            except Exception:
                continue
            if 1 <= d <= 31:
                tmp.append((j, d))
        if len(tmp) >= 8:  # разумный минимум для строки дат
            header_row = r
            day_cols = tmp
            break
    if not day_cols:
        raise ValueError("Не найдена строка с датами (1..31).")

    # --- 1b) если в шапке все дни одинаковые/сломанные — назначаем по порядку ---
    days_in_month = monthrange(year, month)[1]
    uniq = {d for _, d in day_cols}
    if len(uniq) <= max(1, len(day_cols)//4):  # грубый критерий «подозрительно мало»
        # берём только первые N столбцов = число дней в месяце
        day_cols = [(col_idx, i+1) for i, (col_idx, _) in enumerate(day_cols[:days_in_month])]

    # на всякий случай отфильтруем дни > фактических
    day_cols = [(c, d) for (c, d) in day_cols if 1 <= d <= days_in_month]
    if not day_cols:
        raise ValueError("В шапке нет корректных чисел дней для указанного месяца.")

    out = []

    # --- 2) разбираем строки сотрудников ---
    for row in ws.iter_rows(min_row=header_row+1):
        name_cell = row[0]
        full_name = (name_cell.value or "").strip() if name_cell.value else ""
        if not full_name:
            continue
        low = full_name.lower()
        if low.startswith(("plan", "braki", "nazwisko")):
            continue

        for col_idx, day in day_cols:
            cell = ws.cell(row=name_cell.row, column=col_idx)

            # нормализуем код смены
            def _normalize_code(raw):
                if raw is None: return None
                s = str(raw).strip().upper().replace(' ', '')
                if not s or s in {'X','—','-'}: return None
                if s in {'1','2'}: return s
                if s in {'1/B','1B'}: return '1/B'
                if s in {'2/B','2B'}: return '2/B'
                return None

            code = _normalize_code(cell.value)
            if not code:
                continue

            # цвет цифры (font.color) -> lounge
            font_rgb = _rgb_from_openpyxl(getattr(cell.font, 'color', None))
            lounge = 'mazurek' if _is_blue(font_rgb) else ('polonez' if _is_black(font_rgb) else None)

            # цвет заливки (fill.start_color) -> координатор
            fill_rgb = _rgb_from_openpyxl(getattr(cell.fill, 'start_color', None))
            coord_lounge = 'mazurek' if _is_blue(fill_rgb) else ('polonez' if _is_yellow(fill_rgb) else None)

            out.append({
                'name': full_name,
                'date': _date_cls(year, month, day).isoformat(),
                'shift': code,
                'lounge': lounge,
                'coord_lounge': coord_lounge
            })

    return out


@app.post('/api/upload-xlsx')
@jwt_required()
def upload_xlsx():
    """Импорт графика из XLSX (цвет цифры -> lounge, цвет заливки -> coord_lounge)."""

    claims = get_jwt() or {}
    if (claims.get('role') or '').lower() != 'admin':
        return jsonify({'error': 'Tylko administrator może przesyłać XLSX.'}), 403

    # --- Проверка формы ---
    f = request.files.get('file')
    year = int(request.form.get('year') or 0)
    month = int(request.form.get('month') or 0)

    if not f or f.filename == '':
        return jsonify({'error': 'Nie znaleziono pliku (file).'}), 400
    if not (2000 <= year <= 2100 and 1 <= month <= 12):
        return jsonify({'error': 'Podaj poprawny rok i miesiąc.'}), 400

    data = f.read()

    # --- Парсинг XLSX ---
    try:
        rows = parse_schedule_xlsx(data, year, month)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(tb)
        return jsonify({'error': f'Błąd odczytu XLSX: {e}'}), 400

    # --- Очистка данных выбранного месяца ---
    first = _date(year, month, 1)
    last = _date(year, month, monthrange(year, month)[1])

    try:
        db.session.execute(sqltext("""
            DELETE FROM market_offers
            WHERE shift_id IN (
              SELECT id FROM shifts
              WHERE shift_date BETWEEN :d1 AND :d2
            )
        """), {'d1': first, 'd2': last})

        db.session.execute(sqltext("""
            DELETE FROM shifts
            WHERE shift_date BETWEEN :d1 AND :d2
        """), {'d1': first, 'd2': last})

        db.session.flush()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Błąd czyszczenia bazy danych: {e}'}), 500

    # --- Подготовка пользователей ---
    users_by_key = {_norm(u.full_name): u for u in User.query.all()}
    imported = 0
    created_users = []
    seen_pairs = set()  # (user_id, date)

    for r in rows:
        name = (r.get('name') or '').strip()
        if not name:
            continue

        key = _norm(name)
        u = users_by_key.get(key)
        if not u:
            u = User(full_name=name, role='user')
            db.session.add(u)
            db.session.flush()
            users_by_key[key] = u
            created_users.append(u.full_name)

        d_iso = r.get('date')
        if not d_iso:
            continue

        pair = (u.id, d_iso)
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)

        # --- создаём смену ---
        sh = Shift(
            user_id=u.id,
            shift_date=d_iso,
            shift_code=r.get('shift') or '',
            hours=None,
            lounge=(r.get('lounge') or None),
            coord_lounge=(r.get('coord_lounge') or None)
        )

        db.session.add(sh)
        imported += 1

    # --- Сохранение ---
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': f'Błąd zapisu do bazy: {e}'}), 500

    return jsonify({
        'status': 'OK',
        'imported': imported,
        'created_users': created_users
    })


@app.post('/api/upload-text')
@jwt_required()
def upload_text():
    claims = get_jwt() or {}
    if (claims.get('role') or '').lower() != 'admin':
        return jsonify({'error': 'Tylko administrator może przesyłać tekst.'}), 403

    data = request.get_json(force=True) or {}
    text  = (data.get('text') or '').strip()
    month = int(data.get('month') or 0)
    year  = int(data.get('year') or 0)
    if not text:
        return jsonify({'error': 'Pusty tekst.'}), 400
    if not (1 <= month <= 12) or year < 2000:
        return jsonify({'error': 'Podaj poprawny miesiąc i rok.'}), 400

    try:
        rows = _parse_table_no_dates(text, year, month)  # -> list[dict{name, day, code}]
    except Exception as e:
        return jsonify({'error': f'Błąd parsowania: {e}'}), 400

    # диапазон месяца
    first_day = _date(year, month, 1)
    last_day  = _date(year, month, monthrange(year, month)[1])

    # подчистка только выбранного месяца (и зависимых market_offers)
    shift_ids = [s.id for s in Shift.query.filter(
        Shift.shift_date >= first_day,
        Shift.shift_date <= last_day
    ).all()]

    if shift_ids:
        MarketOffer.query.filter(MarketOffer.shift_id.in_(shift_ids)) \
            .delete(synchronize_session=False)
        Shift.query.filter(
            Shift.shift_date >= first_day,
            Shift.shift_date <= last_day
        ).delete(synchronize_session=False)

    # пользователи
    users_by_key = {_norm(u.full_name): u for u in User.query.all()}

    imported = 0
    created_users = []

    for r in rows:
        name = r['name']
        key  = _norm(name)
        user = users_by_key.get(key)
        if not user:
            user = User(full_name=name, role='user')
            db.session.add(user); db.session.flush()
            users_by_key[key] = user
            created_users.append(user.full_name)

        d = _date(year, month, r['day']).isoformat()
        shift_code = r['code']  # '1','2','1/B','2/B', ...
        # часы можешь ставить по умолчанию 9.5, если нужно:
        hours = None

        db.session.add(Shift(
            user_id=user.id,
            shift_date=d,
            shift_code=shift_code,
            hours=hours
        ))
        imported += 1

    db.session.commit()
    return jsonify({'imported': imported, 'created_users': created_users})













# ===== Удаление события =====
@app.route('/api/control/delete', methods=['POST'])
@jwt_required()
def control_delete():
    data = request.get_json()
    event_id = data.get('id')
    reason = data.get('reason', '').strip()
    user_id = get_jwt_identity()

    if not event_id or not reason:
        return jsonify({'error': 'Missing id or reason'}), 400

    # 1️⃣ — достаём данные события
    event_data = db.session.execute(text("""
        SELECT 
            user_id, 
            kind, 
            event_date, 
            hours, 
            time_from, 
            time_to, 
            reason AS event_reason
        FROM control_events
        WHERE id = :eid
    """), {'eid': event_id}).mappings().first()

    if not event_data:
        return jsonify({'error': 'Event not found'}), 404

    # 2️⃣ — копируем в backup
    db.session.execute(text("""
        INSERT INTO control_events_backup (
            event_id, 
            user_id, 
            kind, 
            event_date, 
            hours, 
            time_from, 
            time_to, 
            delay_minutes, 
            reason, 
            deleted_by, 
            deleted_at
        ) VALUES (
            :eid, 
            :uid, 
            :kind, 
            :event_date, 
            :hours, 
            :tf, 
            :tt, 
            :delay, 
            :reason, 
            :deleted_by, 
            CURRENT_TIMESTAMP
        )
    """), {
        'eid': event_id,
        'uid': event_data['user_id'],
        'kind': event_data['kind'],
        'event_date': event_data['event_date'],
        'hours': float(event_data['hours'] or 0),
        'tf': event_data['time_from'],
        'tt': event_data['time_to'],
        'delay': 0,  # можно добавить позже, если появится колонка
        'reason': event_data.get('event_reason') or '',
        'deleted_by': user_id
    })

    # 3️⃣ — логируем удаление
    db.session.execute(text("""
        INSERT INTO control_deleted (event_id, deleted_by, reason, deleted_at)
        VALUES (:eid, :uid, :reason, CURRENT_DATE)
    """), {'eid': event_id, 'uid': user_id, 'reason': reason})

    # 4️⃣ — удаляем из основной таблицы
    db.session.execute(text("DELETE FROM control_events WHERE id = :eid"), {'eid': event_id})
    db.session.commit()

    return jsonify({'status': 'ok'})


# ===== История удалений (список) =====
@app.route('/api/control/deleted', methods=['GET'])
@jwt_required()
def control_deleted_list():
    rows = db.session.execute(text("""
        SELECT 
            c.event_id,
            TO_CHAR(c.deleted_at, 'YYYY-MM-DD') AS deleted_date,
            COALESCE(u.full_name, 'Użytkownik #' || c.deleted_by) AS user_name,
            c.reason
        FROM control_deleted c
        LEFT JOIN users u ON CAST(c.deleted_by AS INTEGER) = u.id
        ORDER BY c.deleted_at DESC
    """)).mappings().all()

    result = []
    for r in rows:
        result.append({
            'event_id': r['event_id'],
            'deleted_date': f"{r['deleted_date']}T00:00:00" if r['deleted_date'] else None,

            'user_name': r['user_name'],
            'reason': r['reason']
        })

    return jsonify(result)


# ===== Детали удалённого события =====
@app.route('/api/control/deleted/<int:event_id>', methods=['GET'])
@jwt_required()
def control_deleted_details(event_id):
    row = db.session.execute(text("""
        SELECT
            ceb.event_id,
            ceb.kind,
            ceb.event_date,
            ceb.time_from,
            ceb.time_to,
            ceb.hours,
            u1.full_name AS user_name,
            u2.full_name AS deleted_by_name,
            ceb.reason,
            ceb.deleted_at
        FROM control_events_backup ceb
        LEFT JOIN users u1 ON u1.id = ceb.user_id
        LEFT JOIN users u2 ON u2.id = ceb.deleted_by
        WHERE ceb.event_id = :eid
    """), {'eid': event_id}).mappings().first()

    if not row:
        return jsonify({'error': 'Not found'}), 404

    return jsonify({
        'event_id': row['event_id'],
        'kind': row['kind'],
        'event_date': row['event_date'].strftime('%Y-%m-%d') if row['event_date'] else None,
        'time_from': row['time_from'],
        'time_to': row['time_to'],
        'hours': float(row['hours']) if row['hours'] else None,
        'user_name': row['user_name'],
        'deleted_by_name': row['deleted_by_name'],
        'deleted_date': row['deleted_at'].strftime('%Y-%m-%d') if row['deleted_at'] else None,
        'reason': row['reason']
    })






# ==== ДОБАВИТЬ ВНИЗ ФАЙЛА server.py (перед if __name__ == '__main__') ====

class CoordShiftReport(db.Model):
    __tablename__ = 'coord_shift_reports'
    id           = db.Column(db.Integer, primary_key=True)
    lounge       = db.Column(db.String(16), nullable=False)  # 'mazurek' | 'polonez'
    shift_type   = db.Column(db.String(16), nullable=False)  # 'morning' | 'evening'
    shift_date   = db.Column(db.Date, nullable=False)
    coord_name   = db.Column(db.String, nullable=False)
    times        = db.Column(db.JSON, default={})
    bars         = db.Column(db.JSON, default={})
    notes        = db.Column(db.JSON, default={})
    created_at   = db.Column(db.DateTime(timezone=True), server_default=func.now())

    def to_dict(self):
        return {
            'id': self.id,
            'lounge': self.lounge,
            'shift_type': self.shift_type,
            'shift_date': self.shift_date.isoformat(),
            'coord_name': self.coord_name,
            'times': self.times,
            'bars': self.bars,
            'notes': self.notes,
            'created_at': self.created_at.isoformat()
        }

# Страница
@app.get('/coord-panel')
@jwt_required()
def coord_panel_page():
    role = (get_jwt().get('role') or '').lower()
    if role != 'coordinator':
        return redirect('/dashboard', 302)
    return render_template('coord-panel.html')

# GET API
@app.get('/api/coord-panel/report')
@jwt_required()
def api_coord_panel_get():
    role = (get_jwt().get('role') or '').lower()
    if role != 'coordinator':
        return jsonify({'error': 'Forbidden'}), 403

    lounge = request.args.get('lounge')
    shift_type = request.args.get('shift_type')
    date_str = request.args.get('date')
    if not all([lounge, shift_type, date_str]):
        return jsonify({'error': 'Missing params'}), 400
    try:
        d = datetime.fromisoformat(date_str).date()
    except:
        return jsonify({'error': 'Bad date'}), 400

    rep = CoordShiftReport.query.filter_by(
        lounge=lounge,
        shift_type=shift_type,
        shift_date=d
    ).first()
    return jsonify(rep.to_dict() if rep else {})

# POST API
@app.post('/api/coord-panel/report')
@jwt_required()
def api_coord_panel_save():
    role = (get_jwt().get('role') or '').lower()
    if role != 'coordinator':
        return jsonify({'error': 'Forbidden'}), 403

    uid = int(get_jwt()['sub'])
    user = User.query.get(uid)
    data = request.get_json(force=True)

    lounge = data['lounge']
    shift_type = data['shift_type']
    shift_date = datetime.fromisoformat(data['shift_date']).date()
    bars = data['bars']
    times = data['times']
    notes = data['notes']

    rep = CoordShiftReport.query.filter_by(
        lounge=lounge,
        shift_type=shift_type,
        shift_date=shift_date
    ).first()

    if not rep:
        rep = CoordShiftReport(
            lounge=lounge,
            shift_type=shift_type,
            shift_date=shift_date,
            coord_name=user.full_name
        )
        db.session.add(rep)

    rep.bars  = bars
    rep.times = times
    rep.notes = notes
    db.session.commit()
    return jsonify({'ok': True})






@app.route('/favicon.ico')
def favicon():
    # Возвращаем 204 No Content, чтобы браузер отстал
    return '', 204


# ---------------------------------
# Pages
# ---------------------------------
@app.get('/')
def index_page():
    return render_template('index.html')

@app.get('/dashboard')
def dashboard_page():
    return render_template('dashboard.html')


@app.get('/admin')  # НИКАКИХ @jwt_required тут быть не должно
def admin_page():
    # отдаем статическую страницу; права проверит admin.js
    return send_file(os.path.join(app.root_path, 'templates', 'admin.html'))



@app.get('/stats')
def stats_page():
    return render_template('stats.html')

@app.get('/api/health')
def health():
    return jsonify({'ok': True, 'ts': datetime.utcnow().isoformat()})

@app.get('/start')
def start_page():
    return render_template('start.html')

@app.get('/proposals')
def proposals_page():
    return render_template('proposals.html')

@app.get('/market')
def market_page():
    return render_template('market.html')

@app.get('/control')
def control_page():
    return render_template('control.html')




@app.errorhandler(Exception)
def catch_all(e):
    app.logger.exception("Unhandled")
    return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    with app.app_context():
        ensure_coord_lounge_column()
        ensure_lounge_column()   # ← ВАЖНО
    app.run(host='0.0.0.0', port=port, debug=True, use_reloader=False)




























